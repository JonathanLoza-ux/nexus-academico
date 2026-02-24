"""Servicios de exportacion para reportes administrativos."""

import csv
from datetime import datetime
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def build_xlsx_response(*, filename_base: str, sheet_name: str, headers: list, rows: list, response_cls):
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Reporte")[:31]

    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    header_fill = PatternFill(start_color="13233F", end_color="13233F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center

    for idx, _ in enumerate(headers, start=1):
        max_len = len(str(headers[idx - 1]))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=idx).value
            val_len = len(str(val)) if val is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 52)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"{filename_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return response_cls(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def build_users_export_rows(*, users, now_utc, user_status_data_fn, format_dt_human_fn):
    values = []
    for user in users:
        status_key, status_label, suspended_until = user_status_data_fn(user, now_utc)
        values.append([
            user.id,
            user.name or "",
            user.email or "",
            user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else "",
            status_label,
            format_dt_human_fn(suspended_until) if status_key == "suspendida" else "",
            int(user.chat_count or 0),
            int(user.message_count or 0),
        ])
    return values


def build_login_attempt_export_rows(login_attempt_rows):
    values = []
    for row in login_attempt_rows:
        values.append([
            row.id,
            row.ip or "",
            row.email or "",
            int(row.attempts or 0),
            row.first_attempt_at.strftime('%Y-%m-%d %H:%M:%S') if row.first_attempt_at else "",
            row.blocked_until.strftime('%Y-%m-%d %H:%M:%S') if row.blocked_until else "",
        ])
    return values


def build_audit_export_rows(*, enriched_rows, mask_sensitive_text_fn):
    values = []
    for item in enriched_rows:
        log_row = item["log_row"]
        actor_user = item["actor_user"]
        values.append([
            log_row.created_at.strftime('%Y-%m-%d %H:%M:%S') if log_row.created_at else "",
            item["severity_label"],
            actor_user.email if actor_user else "",
            log_row.action or "",
            log_row.target_user_id or "",
            log_row.ip or "",
            item["request_id"] or "",
            item["method"] or "",
            item["path"] or "",
            item["module_name"] or "",
            mask_sensitive_text_fn(log_row.detail or ""),
        ])
    return values


def build_audit_csv_response(*, enriched_rows, mask_sensitive_text_fn, response_cls):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Fecha", "Severidad", "Actor", "Accion", "Usuario objetivo", "IP", "Request ID", "Metodo", "Ruta", "Modulo", "Detalle"])
    for item in enriched_rows:
        log_row = item["log_row"]
        actor_user = item["actor_user"]
        writer.writerow([
            log_row.created_at.strftime('%Y-%m-%d %H:%M:%S') if log_row.created_at else "",
            item["severity_label"],
            actor_user.email if actor_user else "",
            log_row.action or "",
            log_row.target_user_id or "",
            log_row.ip or "",
            item["request_id"] or "",
            item["method"] or "",
            item["path"] or "",
            item["module_name"] or "",
            mask_sensitive_text_fn(log_row.detail or ""),
        ])

    data = output.getvalue()
    output.close()
    return response_cls(
        data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=auditoria_admin.csv"},
    )


def build_audit_pdf_response(*, enriched_rows, mask_sensitive_text_fn, response_cls):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except Exception as exc:
        raise RuntimeError("reportlab_missing") from exc

    def _short(v, size):
        txt = str(v or "")
        return txt if len(txt) <= size else (txt[: size - 3] + "...")

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    y = height - 12 * mm

    c.setTitle("Auditoria administrativa")
    c.setFont("Helvetica-Bold", 12)
    c.drawString(12 * mm, y, "Auditoria administrativa")
    c.setFont("Helvetica", 8)
    c.drawRightString(width - 12 * mm, y, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 7 * mm

    headers = ["Fecha", "Sev", "Accion", "Actor", "IP", "RID", "Detalle"]
    widths = [32, 12, 40, 48, 26, 26, 96]

    def _draw_header(cur_y):
        c.setFont("Helvetica-Bold", 8)
        x = 10 * mm
        for idx, head in enumerate(headers):
            c.drawString(x, cur_y, head)
            x += widths[idx] * mm
        return cur_y - 4.5 * mm

    y = _draw_header(y)
    c.setFont("Helvetica", 7)
    for item in enriched_rows:
        log_row = item["log_row"]
        actor_user = item["actor_user"]
        row_vals = [
            log_row.created_at.strftime('%Y-%m-%d %H:%M:%S') if log_row.created_at else "-",
            item["severity_label"],
            log_row.action or "-",
            actor_user.email if actor_user else "-",
            log_row.ip or "-",
            item["request_id"] or "-",
            mask_sensitive_text_fn(log_row.detail or "-"),
        ]
        if y < 14 * mm:
            c.showPage()
            y = height - 12 * mm
            y = _draw_header(y)
            c.setFont("Helvetica", 7)
        x = 10 * mm
        for idx, val in enumerate(row_vals):
            limit = 18 if idx in (0, 2, 3, 6) else 12
            c.drawString(x, y, _short(val, limit))
            x += widths[idx] * mm
        y -= 4.2 * mm

    c.save()
    buffer.seek(0)
    return response_cls(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=auditoria_admin.pdf"},
    )


def build_users_json_payload(*, users, now_utc, user_status_data_fn, format_dt_human_fn):
    export_rows = []
    for user in users:
        status_key, status_label, suspended_until = user_status_data_fn(user, now_utc)
        export_rows.append([
            user.id,
            user.name or "",
            user.email or "",
            user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else "-",
            status_label,
            format_dt_human_fn(suspended_until) if status_key == "suspendida" else "-",
            int(user.chat_count or 0),
            int(user.message_count or 0),
        ])

    return {
        "success": True,
        "title": "Usuarios Nexus",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sections": [
            {
                "title": "Usuarios",
                "columns": ["ID", "Nombre", "Email", "Registro", "Estado", "Suspendida hasta", "Chats", "Mensajes"],
                "rows": export_rows,
            }
        ],
    }


def build_audit_json_payload(*, enriched_rows, mask_sensitive_text_fn):
    return {
        "success": True,
        "title": "Auditoria administrativa",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sections": [
            {
                "title": "Eventos de auditoria",
                "columns": ["Fecha", "Severidad", "Actor", "Accion", "Usuario objetivo", "IP", "Request ID", "Metodo", "Ruta", "Modulo", "Detalle"],
                "rows": [
                    [
                        item["log_row"].created_at.strftime('%Y-%m-%d %H:%M:%S') if item["log_row"].created_at else "-",
                        item["severity_label"],
                        item["actor_user"].email if item["actor_user"] else "-",
                        item["log_row"].action or "-",
                        item["log_row"].target_user_id or "-",
                        item["log_row"].ip or "-",
                        item["request_id"] or "-",
                        item["method"] or "-",
                        item["path"] or "-",
                        item["module_name"] or "-",
                        mask_sensitive_text_fn(item["log_row"].detail or "-"),
                    ]
                    for item in enriched_rows
                ],
            }
        ],
    }


def build_security_json_payload(*, login_rows, reset_rows, rate_rows):
    return {
        "success": True,
        "title": "Reporte de seguridad",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sections": [
            {
                "title": "Intentos de login",
                "columns": ["ID", "IP", "Email", "Intentos", "Primer intento", "Bloqueado hasta"],
                "rows": [
                    [
                        r.id,
                        r.ip or "-",
                        r.email or "-",
                        r.attempts or 0,
                        r.first_attempt_at.strftime('%Y-%m-%d %H:%M:%S') if r.first_attempt_at else "-",
                        r.blocked_until.strftime('%Y-%m-%d %H:%M:%S') if r.blocked_until else "-",
                    ]
                    for r in login_rows
                ],
            },
            {
                "title": "Reset por IP",
                "columns": ["IP", "Intentos", "Primer intento", "Bloqueado hasta"],
                "rows": [
                    [
                        r.ip or "-",
                        r.attempts or 0,
                        r.first_attempt_at.strftime('%Y-%m-%d %H:%M:%S') if r.first_attempt_at else "-",
                        r.blocked_until.strftime('%Y-%m-%d %H:%M:%S') if r.blocked_until else "-",
                    ]
                    for r in reset_rows
                ],
            },
            {
                "title": "Rate limits",
                "columns": ["Key", "Count", "Ventana", "Bloqueado hasta"],
                "rows": [
                    [
                        r.key or "-",
                        r.count or 0,
                        r.window_start.strftime('%Y-%m-%d %H:%M:%S') if r.window_start else "-",
                        r.blocked_until.strftime('%Y-%m-%d %H:%M:%S') if r.blocked_until else "-",
                    ]
                    for r in rate_rows
                ],
            },
        ],
    }
