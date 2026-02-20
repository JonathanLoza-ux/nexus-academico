import os
import csv
import random
import re
import json
from datetime import datetime, timedelta, timezone  # ✅ Cambio 1: Añadido timezone
from io import BytesIO, StringIO
import requests
import logging
import time
import uuid
from collections import Counter
from functools import wraps
import math
from urllib.parse import urlencode, urlparse

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, abort, Response
from dotenv import load_dotenv

import google.generativeai as genai
from google.api_core import exceptions as gexc
from PIL import Image

from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from sqlalchemy import inspect, text, func
from sqlalchemy.exc import OperationalError

from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

import cloudinary
import cloudinary.uploader

from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from flask_mail import Mail, Message as MailMessage

from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================================================
# 1) CARGA DE VARIABLES DE ENTORNO (.env)
# =========================================================
load_dotenv()

# Cloudinary (para imágenes)
CLOUDINARY_URL = os.getenv("CLOUDINARY_URL")
if CLOUDINARY_URL:
    # FIX: recargar config desde env para parsear api_key/api_secret
    # (evita el error "Must supply api_key" al subir imagen)
    cloudinary.reset_config()
    cloudinary.config(secure=True)
    # DEBUG (DEV): confirma que Cloudinary cargó credenciales sin exponer secretos
    if (os.getenv("ENVIRONMENT") or "dev").strip().lower() == "dev":
        cfg = cloudinary.config()
        api_key = cfg.api_key or ""
        api_key_mask = f"...{api_key[-4:]}" if api_key else "None"
        print("=== CLOUDINARY DEBUG ===")
        print("CLOUDINARY cloud_name:", repr(cfg.cloud_name))
        print("CLOUDINARY api_key:", repr(api_key_mask))
        print("========================")

# Gemini Keys (varias claves separadas por coma)
claves_string = os.getenv("GEMINI_KEYS")
if not claves_string:
    print("⚠️ ADVERTENCIA: No se encontró 'GEMINI_KEYS' en el .env.")
    LISTA_DE_CLAVES = []
else:
    LISTA_DE_CLAVES = [key.strip() for key in claves_string.split(',') if key.strip()]

# =========================================================
# PASO B: Modificar configurar_gemini_random() para que devuelva la key
# =========================================================
def configurar_gemini_random():
    """Elige una clave random para Gemini y la configura."""
    if not LISTA_DE_CLAVES:
        return None
    clave_elegida = random.choice(LISTA_DE_CLAVES)
    genai.configure(api_key=clave_elegida)
    return clave_elegida


configurar_gemini_random()

# =========================================================
# 2) CONFIGURACIÓN DE LA APP
# =========================================================
app = Flask(__name__)

# =========================================================
# LOGS PRO (Fase 1.9)
# =========================================================
logger = logging.getLogger("nexus")
logger.setLevel(logging.INFO)

if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)

APP_STARTED_AT = datetime.now(timezone.utc).replace(tzinfo=None)
ADMIN_LOG_RETENTION_DAYS = max(30, int((os.getenv("ADMIN_LOG_RETENTION_DAYS") or "180").strip()))
ADMIN_LOG_CLEANUP_INTERVAL_S = max(600, int((os.getenv("ADMIN_LOG_CLEANUP_INTERVAL_S") or "21600").strip()))
_ADMIN_LOG_CLEANUP_LAST_TS = 0.0

# ==========================
# 🔐 Seguridad base (Fase 0)
# ==========================

# Entorno: dev / prod
ENVIRONMENT = (os.getenv("ENVIRONMENT") or "dev").strip().lower()

# Panel admin:
# Puedes definir varios correos separados por coma en SUPER_ADMIN_EMAILS.
_super_admin_env = (os.getenv("SUPER_ADMIN_EMAILS") or "jonathandavidloza@gmail.com").strip()
SUPER_ADMIN_EMAILS = {
    e.strip().lower() for e in _super_admin_env.split(",") if e.strip()
}

ALL_ADMIN_PERMISSIONS = {
    "view_dashboard",
    "view_users",
    "view_conversations",
    "view_logs",
    "view_security",
    "export_reports",
    "manage_users",
    "manage_admins",
    "manage_settings",
}

DEFAULT_ADMIN_PERMISSIONS = {
    "view_dashboard",
    "view_users",
    "view_conversations",
    "view_logs",
    "export_reports",
}

PERMISSION_LABELS_ES = {
    "view_dashboard": "Ver panel general",
    "view_users": "Ver usuarios",
    "view_conversations": "Ver conversaciones",
    "view_logs": "Ver registros (logs)",
    "view_security": "Ver seguridad",
    "export_reports": "Exportar reportes",
    "manage_users": "Gestionar usuarios",
    "manage_admins": "Gestionar administradores",
    "manage_settings": "Gestionar configuracion",
}

PERMISSION_GROUPS_ES = [
    {
        "title": "Panel y reportes",
        "icon": "fa-chart-line",
        "codes": ["view_dashboard", "export_reports"],
    },
    {
        "title": "Usuarios y conversaciones",
        "icon": "fa-users",
        "codes": ["view_users", "view_conversations", "manage_users"],
    },
    {
        "title": "Seguridad y registros",
        "icon": "fa-shield-halved",
        "codes": ["view_security", "view_logs"],
    },
    {
        "title": "Administracion",
        "icon": "fa-user-shield",
        "codes": ["manage_admins", "manage_settings"],
    },
]

# Secret key desde entorno (MUY IMPORTANTE en producción)
app.secret_key = (os.getenv("SECRET_KEY") or "dev_secret_key_change_me").strip()

# Cookies de sesión más seguras:
# - SECURE solo cuando estés en HTTPS (Render sí, local no)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = (ENVIRONMENT == "prod")  # True en Render, False local

# Opcional: duración de sesión (ej. 7 días)
# from datetime import timedelta
# app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)

# ✅ ProxyFix: permite obtener IP real y scheme correcto detrás de Render
# x_for=1 y x_proto=1 suelen ser suficientes en Render
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

# ✅ Para que url_for(..., _external=True) genere bien enlaces en producción (Render/otro)
if os.getenv("SERVER_NAME"):
    app.config["SERVER_NAME"] = os.getenv("SERVER_NAME").strip()

# Forzar HTTPS en producción
if ENVIRONMENT == "prod":
    app.config["PREFERRED_URL_SCHEME"] = "https"

# ✅ Serializador para tokens de reset
serializer = URLSafeTimedSerializer(app.secret_key)
RESET_TOKEN_MAX_AGE = 20 * 60  # 20 minutos

# ✅ Modo reset:
# - dev  -> imprime link en consola
# - smtp -> envía correo real por SMTP (Brevo / Gmail / etc)
# - brevo_api -> usa la API REST de Brevo
RESET_MODE = (os.getenv("RESET_MODE") or "dev").strip().lower()
SUPPORT_WHATSAPP = (os.getenv("SUPPORT_WHATSAPP") or "50364254348").strip().replace("+", "")

# 3) Agregar variables Brevo después de RESET_MODE
BREVO_API_KEY = (os.getenv("BREVO_API_KEY") or "").strip()
BREVO_SENDER_NAME = (os.getenv("BREVO_SENDER_NAME") or "Nexus Academy").strip()
BREVO_SENDER_EMAIL = (os.getenv("BREVO_SENDER_EMAIL") or "").strip()

# Si falta api key o sender y estás en modo brevo_api, cae a dev
if RESET_MODE == "brevo_api":
    if (not BREVO_API_KEY) or (not BREVO_SENDER_EMAIL):
        print("⚠️ RESET_MODE=brevo_api pero falta BREVO_API_KEY o BREVO_SENDER_EMAIL. Forzando RESET_MODE=dev")
        RESET_MODE = "dev"

# 🔒 Bloque repetido (comentado) - NO BORRAR, solo dejarlo desactivado
# from werkzeug.middleware.proxy_fix import ProxyFix
# app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
# app.config["PREFERRED_URL_SCHEME"] = "https"

# =========================================================
# 3) CONFIG SMTP (BREVO / GMAIL / CUALQUIERA)
# =========================================================
MAIL_SERVER = (os.getenv("MAIL_SERVER") or "").strip()
MAIL_PORT = int((os.getenv("MAIL_PORT") or "587").strip())
MAIL_USE_TLS = (os.getenv("MAIL_USE_TLS") or "1").strip() == "1"
MAIL_USERNAME = (os.getenv("MAIL_USERNAME") or "").strip()
MAIL_PASSWORD = (os.getenv("MAIL_PASSWORD") or "").strip()
MAIL_DEFAULT_SENDER = (os.getenv("MAIL_DEFAULT_SENDER") or MAIL_USERNAME).strip()

app.config["MAIL_SERVER"] = MAIL_SERVER
app.config["MAIL_PORT"] = MAIL_PORT
app.config["MAIL_USE_TLS"] = MAIL_USE_TLS
app.config["MAIL_USERNAME"] = MAIL_USERNAME
app.config["MAIL_PASSWORD"] = MAIL_PASSWORD
app.config["MAIL_DEFAULT_SENDER"] = MAIL_DEFAULT_SENDER

# ✅ Timeout para evitar cuelgues en servidor
app.config["MAIL_TIMEOUT"] = int((os.getenv("MAIL_TIMEOUT") or "10").strip())

# ✅ FAIL-SAFE: si dicen smtp pero falta config, fuerza dev para que no intente localhost
if RESET_MODE == "smtp":
    if not MAIL_SERVER or not MAIL_USERNAME or not MAIL_PASSWORD:
        print("⚠️ SMTP incompleto. Forzando RESET_MODE=dev para evitar errores.")
        RESET_MODE = "dev"

# ✅ DEBUG (NO imprime password)
print("=== SMTP DEBUG ===")
print("MAIL_SERVER:", repr(app.config.get("MAIL_SERVER")))
print("MAIL_PORT:", repr(app.config.get("MAIL_PORT")))
print("MAIL_USE_TLS:", repr(app.config.get("MAIL_USE_TLS")))
print("MAIL_USERNAME:", repr(app.config.get("MAIL_USERNAME")))
print("MAIL_DEFAULT_SENDER:", repr(app.config.get("MAIL_DEFAULT_SENDER")))
print("RESET_MODE:", repr(RESET_MODE))
print("==================")

mail = Mail(app)

# =========================================================
# 4) BASE DE DATOS (Clever Cloud MySQL)
# =========================================================
# ✅ Ahora se lee desde .env (más seguro)
uri_db = (os.getenv("DATABASE_URL") or "").strip()

if not uri_db:
    raise RuntimeError("❌ Falta DATABASE_URL en el .env / Render Environment Variables")

# Compatibilidad: SQLAlchemy 2 usa MySQLdb por defecto con mysql://
# y en este proyecto usamos PyMySQL.
if uri_db.startswith("mysql://"):
    uri_db = "mysql+pymysql://" + uri_db[len("mysql://"):]

app.config['SQLALCHEMY_DATABASE_URI'] = uri_db
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 280,
    'connect_args': {'connect_timeout': 10}
}

db = SQLAlchemy(app)

# =========================================================
# 5) LOGIN (Flask-Login)
# =========================================================
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login_page'
login_manager.login_message = None

# =========================================================
# 6) VALIDACIONES Y SUBIDAS
# =========================================================
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PASSWORD_RE = re.compile(r"^(?=.*[A-Za-z])(?=.*[^A-Za-z0-9]).{6,}$")

# =========================================================
# Fase 1 — Paso 2: Helper de IP real (lo usaremos en reset/login)
# =========================================================
def get_client_ip():
    """
    Obtiene la dirección IP real del cliente considerando proxies intermedios.
    """
    if request.access_route:
        return request.access_route[0]
    return request.remote_addr or "0.0.0.0"


@app.before_request
def _before_request_logging():
    request._start_time = time.time()
    request._rid = uuid.uuid4().hex[:12]
    _cleanup_old_admin_logs(force=False)


@app.before_request
def _before_request_enforce_active_account():
    # Evita romper carga de recursos estáticos
    if request.path.startswith("/static/"):
        return None

    if not current_user.is_authenticated:
        return None

    email_hint = (getattr(current_user, "email", "") or "").strip().lower()
    user_id = getattr(current_user, "id", None)
    suspension_until = _active_suspension_until(current_user)

    if bool(getattr(current_user, "is_active_account", True)) and not suspension_until:
        # Cierre forzado de sesion si un admin lo marco desde panel de seguridad.
        ctl = UserSessionControl.query.filter_by(user_id=user_id).first() if user_id else None
        force_after = to_naive_utc(getattr(ctl, "force_logout_after", None))
        if force_after:
            login_at_ts = int(session.get("login_at_ts") or 0)
            force_ts = int(force_after.replace(tzinfo=timezone.utc).timestamp())
            if login_at_ts <= 0 or login_at_ts < force_ts:
                logout_user()
                for key in ["_user_id", "_fresh", "_id", "remember_token", "login_at_ts"]:
                    session.pop(key, None)
                _set_login_help_mode("support", email_hint=email_hint)
                flash("Tu sesion fue cerrada por seguridad. Inicia sesion nuevamente.", "error")
                log_event("FORCE_LOGOUT_SECURITY", user_id=user_id, email=email_hint, path=request.path)
                return redirect(url_for('login_page'))
        return None

    logout_user()
    for key in ["_user_id", "_fresh", "_id", "remember_token", "login_at_ts"]:
        session.pop(key, None)

    if suspension_until:
        msg = f"Tu cuenta esta suspendida hasta {_format_dt_human(suspension_until)}."
        _set_login_help_mode("support", email_hint=email_hint)
        flash(msg + " Contacta al administrador.", "error")
        log_event(
            "FORCE_LOGOUT_SUSPENDED",
            user_id=user_id,
            email=email_hint,
            path=request.path,
            suspended_until=_format_dt_human(suspension_until),
        )
    else:
        _set_login_help_mode("support", email_hint=email_hint)
        flash("Tu cuenta esta desactivada. Contacta al administrador.", "error")
        log_event("FORCE_LOGOUT_INACTIVE", user_id=user_id, email=email_hint, path=request.path)

    redirect_url = url_for('login_page')
    wants_json = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or request.path.startswith("/chat")
        or request.path.startswith("/shared_send")
        or request.path.startswith("/shared_regenerate")
        or request.path.startswith("/feedback")
    )
    if wants_json:
        err_msg = (
            f"Tu cuenta esta suspendida hasta {_format_dt_human(suspension_until)}."
            if suspension_until else
            "Tu cuenta esta desactivada. Contacta al administrador."
        )
        return jsonify({
            "success": False,
            "error": err_msg,
            "redirect": redirect_url,
        }), 401

    return redirect(redirect_url)


@app.after_request
def _after_request_logging(response):
    try:
        elapsed_ms = int((time.time() - getattr(request, "_start_time", time.time())) * 1000)
        rid = getattr(request, "_rid", "-")
        ip = get_client_ip()

        # No ensuciar logs con archivos estáticos
        if not request.path.startswith("/static/"):
            logger.info(
                f"HTTP rid={rid} method={request.method} path={request.path} "
                f"status={response.status_code} ip={ip} ms={elapsed_ms}"
            )
    except Exception:
        pass
    return response


def log_event(event: str, **fields):
    """
    Log estructurado tipo:
    EVENT rid=xxxx key=value key=value
    """
    rid = getattr(request, "_rid", "-")
    base = f"{event} rid={rid}"

    parts = []
    for k, v in fields.items():
        if v is None:
            continue
        parts.append(f"{k}={str(v).replace(' ', '_')}")

    msg = base + (" " + " ".join(parts) if parts else "")
    logger.info(msg)


UPLOAD_DIR = os.path.join(app.root_path, 'static', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# =========================================================
# HELPERS DE FECHAS NAIVE UTC (compatibles con MySQL)
# =========================================================
def utcnow_naive():
    """UTC naive sin usar datetime.utcnow() (evita DeprecationWarning)."""
    return datetime.now(timezone.utc).replace(tzinfo=None)

def to_naive_utc(dt):
    """Convierte aware->naive UTC. Si ya es naive, lo deja igual."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


def _format_dt_human(dt):
    dt_n = to_naive_utc(dt)
    if not dt_n:
        return "-"
    return dt_n.strftime('%Y-%m-%d %H:%M:%S')


def _active_suspension_until(user, now_dt=None):
    if not user:
        return None
    now_dt = now_dt or utcnow_naive()
    until = to_naive_utc(getattr(user, "suspended_until", None))
    if until and until > now_dt:
        return until
    return None


def _user_status_data(user, now_dt=None):
    now_dt = now_dt or utcnow_naive()
    is_active = bool(getattr(user, "is_active_account", True))
    until = _active_suspension_until(user, now_dt)
    if not is_active:
        return "desactivada", "Desactivada", None
    if until:
        return "suspendida", "Suspendida", until
    return "activa", "Activa", None

# =========================================================
# 7) MODELOS
# =========================================================
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=utcnow_naive, index=True)
    is_active_account = db.Column(db.Boolean, default=True, nullable=False, index=True)
    suspended_until = db.Column(db.DateTime, nullable=True, index=True)
    conversations = db.relationship('Conversation', backref='owner', lazy=True)
    saved_messages = db.relationship('SavedMessage', backref='owner', lazy=True, cascade="all, delete-orphan")


class Conversation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), default="Nuevo Chat")
    created_at = db.Column(db.DateTime, default=utcnow_naive)  # ✅ Cambiado a naive UTC
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    messages = db.relationship('Message', backref='conversation', lazy=True, cascade="all, delete-orphan")


class Message(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    content = db.Column(db.Text, nullable=False)
    sender = db.Column(db.String(10), nullable=False)
    timestamp = db.Column(db.DateTime, default=utcnow_naive)  # ✅ Cambiado a naive UTC
    has_image = db.Column(db.Boolean, default=False)
    conversation_id = db.Column(db.Integer, db.ForeignKey('conversation.id'), nullable=False)


class SavedMessage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=utcnow_naive, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)


class AdminRole(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, unique=True, index=True)
    role = db.Column(db.String(20), nullable=False, default="admin", index=True)  # admin | super_admin
    permissions_json = db.Column(db.Text, nullable=False, default="[]")
    is_active = db.Column(db.Boolean, nullable=False, default=True, index=True)
    granted_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=utcnow_naive, index=True)
    updated_at = db.Column(db.DateTime, default=utcnow_naive, onupdate=utcnow_naive)


class AdminAuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    actor_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True, index=True)
    action = db.Column(db.String(120), nullable=False, index=True)
    target_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True, index=True)
    detail = db.Column(db.Text, nullable=True)
    ip = db.Column(db.String(64), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=utcnow_naive, index=True)


class SharedConversation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    token = db.Column(db.String(64), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=utcnow_naive)
    conversation_id = db.Column(db.Integer, db.ForeignKey('conversation.id'), nullable=False, index=True)
    owner_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    read_only = db.Column(db.Boolean, default=True)
    allow_export = db.Column(db.Boolean, default=True)
    allow_copy = db.Column(db.Boolean, default=True)
    allow_feedback = db.Column(db.Boolean, default=True)
    allow_regenerate = db.Column(db.Boolean, default=False)
    allow_edit = db.Column(db.Boolean, default=False)


class SharedViewerPresence(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    token = db.Column(db.String(64), index=True, nullable=False)
    email = db.Column(db.String(120), index=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    last_seen = db.Column(db.DateTime, default=utcnow_naive, index=True)


class ResetRequest(db.Model):
    """
    Controla intentos y cooldown para evitar spam de correos.
    """
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), index=True, nullable=False)
    last_sent_at = db.Column(db.DateTime, nullable=True)
    attempts = db.Column(db.Integer, default=0)
    first_attempt_at = db.Column(db.DateTime, nullable=True)


# =========================================================
# Fase 1 — Paso 3: Nuevos modelos (DB) para límites por IP y login attempts
# =========================================================
class ResetIPRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ip = db.Column(db.String(64), index=True, unique=True, nullable=False)

    last_sent_at = db.Column(db.DateTime, nullable=True)
    attempts = db.Column(db.Integer, default=0)
    first_attempt_at = db.Column(db.DateTime, nullable=True)

    blocked_until = db.Column(db.DateTime, nullable=True)


class LoginAttempt(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    ip = db.Column(db.String(64), index=True, nullable=False)
    email = db.Column(db.String(100), index=True, nullable=True)

    attempts = db.Column(db.Integer, default=0)
    first_attempt_at = db.Column(db.DateTime, nullable=True)

    blocked_until = db.Column(db.DateTime, nullable=True)


# =========================================================
# Fase 2 — Paso 2.1: Modelo para Rate Limit genérico (DB)
# =========================================================
class RateLimit(db.Model):
    """
    Tabla genérica para rate limiting persistente en DB.
    Funciona con múltiples workers (Render).
    """
    id = db.Column(db.Integer, primary_key=True)

    key = db.Column(db.String(200), unique=True, nullable=False, index=True)
    window_start = db.Column(db.DateTime, nullable=True)
    count = db.Column(db.Integer, default=0)

    blocked_until = db.Column(db.DateTime, nullable=True)


class SecurityBlock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    block_type = db.Column(db.String(20), nullable=False, index=True)  # email | ip
    target = db.Column(db.String(190), nullable=False, index=True)
    reason = db.Column(db.String(255), nullable=True)
    blocked_until = db.Column(db.DateTime, nullable=False, index=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False, index=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=utcnow_naive, nullable=False, index=True)
    updated_at = db.Column(db.DateTime, default=utcnow_naive, onupdate=utcnow_naive, nullable=False)


class UserSessionControl(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False, index=True)
    force_logout_after = db.Column(db.DateTime, nullable=True, index=True)
    updated_at = db.Column(db.DateTime, default=utcnow_naive, onupdate=utcnow_naive, nullable=False)


def _normalize_email(value: str) -> str:
    return (value or "").strip().lower()


def _normalize_ip(value: str) -> str:
    return (value or "").strip()


def _active_security_block(block_type: str, target: str, now_utc=None):
    now_utc = now_utc or utcnow_naive()
    clean_type = (block_type or "").strip().lower()
    clean_target = _normalize_email(target) if clean_type == "email" else _normalize_ip(target)
    if clean_type not in {"email", "ip"} or not clean_target:
        return None
    return (
        SecurityBlock.query.filter_by(
            block_type=clean_type,
            target=clean_target,
            is_active=True,
        )
        .filter(SecurityBlock.blocked_until > now_utc)
        .order_by(SecurityBlock.blocked_until.desc())
        .first()
    )


def _security_block_wait_seconds(block_row, now_utc=None) -> int:
    now_utc = now_utc or utcnow_naive()
    if not block_row or not block_row.blocked_until:
        return 0
    until = to_naive_utc(block_row.blocked_until)
    if not until or until <= now_utc:
        return 0
    return max(1, int((until - now_utc).total_seconds()))


def _mark_force_logout(user_id: int):
    if not user_id:
        return
    row = UserSessionControl.query.filter_by(user_id=user_id).first()
    now_utc = utcnow_naive()
    if not row:
        row = UserSessionControl(user_id=user_id, force_logout_after=now_utc)
        db.session.add(row)
    else:
        row.force_logout_after = now_utc
        row.updated_at = now_utc
    db.session.commit()


def _loads_permissions(raw: str):
    if not raw:
        return set()
    try:
        data = json.loads(raw)
        if isinstance(data, list):
            return {str(x).strip() for x in data if str(x).strip()}
    except Exception:
        pass
    return set()


def _dumps_permissions(perms):
    clean = sorted({p for p in perms if p in ALL_ADMIN_PERMISSIONS})
    return json.dumps(clean, ensure_ascii=True)


def _permission_label_es(code: str):
    return PERMISSION_LABELS_ES.get(code, code)


def _is_super_admin_email(email: str) -> bool:
    return _normalize_email(email) in SUPER_ADMIN_EMAILS


def _get_admin_role_record(user_id: int):
    if not user_id:
        return None
    return AdminRole.query.filter_by(user_id=user_id, is_active=True).first()


def _effective_admin_role(user: User):
    if not user:
        return None
    if _is_super_admin_email(user.email):
        return "super_admin"
    rec = _get_admin_role_record(user.id)
    if rec and rec.role in ("admin", "super_admin"):
        return rec.role
    return None


def _effective_admin_permissions(user: User):
    role = _effective_admin_role(user)
    if role == "super_admin":
        return set(ALL_ADMIN_PERMISSIONS)
    if role == "admin":
        rec = _get_admin_role_record(user.id)
        if not rec:
            return set(DEFAULT_ADMIN_PERMISSIONS)
        perms = _loads_permissions(rec.permissions_json)
        return perms or set(DEFAULT_ADMIN_PERMISSIONS)
    return set()


def _admin_has_permission(user: User, permission: str):
    if not user:
        return False
    perms = _effective_admin_permissions(user)
    return permission in perms


def _ensure_super_admin_membership(user: User):
    """
    Garantiza que el super admin por correo tenga registro persistente en DB.
    """
    if not user or not _is_super_admin_email(user.email):
        return
    rec = AdminRole.query.filter_by(user_id=user.id).first()
    if not rec:
        rec = AdminRole(
            user_id=user.id,
            role="super_admin",
            permissions_json=_dumps_permissions(ALL_ADMIN_PERMISSIONS),
            is_active=True,
            granted_by_user_id=user.id,
        )
        db.session.add(rec)
        db.session.commit()
        return

    changed = False
    if rec.role != "super_admin":
        rec.role = "super_admin"
        changed = True
    if not rec.is_active:
        rec.is_active = True
        changed = True

    desired = _dumps_permissions(ALL_ADMIN_PERMISSIONS)
    if rec.permissions_json != desired:
        rec.permissions_json = desired
        changed = True

    if changed:
        rec.updated_at = utcnow_naive()
        db.session.commit()


def _bootstrap_super_admin_roles():
    if not SUPER_ADMIN_EMAILS:
        return
    users = User.query.filter(func.lower(User.email).in_(list(SUPER_ADMIN_EMAILS))).all()
    for user in users:
        _ensure_super_admin_membership(user)


def _add_admin_audit(action: str, target_user_id=None, detail=None):
    try:
        detail_parts = []
        if detail:
            detail_parts.append(str(detail).strip())
        if request:
            rid = getattr(request, "_rid", "") or ""
            meta = []
            if rid:
                meta.append(f"rid={rid}")
            if request.method:
                meta.append(f"method={request.method}")
            if request.path:
                meta.append(f"path={request.path}")
            if request.endpoint:
                meta.append(f"endpoint={request.endpoint}")
            if meta:
                detail_parts.append("; ".join(meta))
        row = AdminAuditLog(
            actor_user_id=current_user.id if current_user.is_authenticated else None,
            action=action,
            target_user_id=target_user_id,
            detail=("; ".join([p for p in detail_parts if p]))[:2000],
            ip=get_client_ip() if request else None,
        )
        db.session.add(row)
        db.session.commit()
    except Exception:
        db.session.rollback()


def _cleanup_old_admin_logs(force=False):
    global _ADMIN_LOG_CLEANUP_LAST_TS
    now_ts = time.time()
    if not force and (now_ts - _ADMIN_LOG_CLEANUP_LAST_TS) < ADMIN_LOG_CLEANUP_INTERVAL_S:
        return 0
    cutoff = utcnow_naive() - timedelta(days=ADMIN_LOG_RETENTION_DAYS)
    try:
        deleted = (
            AdminAuditLog.query
            .filter(AdminAuditLog.created_at < cutoff)
            .delete(synchronize_session=False)
        )
        db.session.commit()
        _ADMIN_LOG_CLEANUP_LAST_TS = now_ts
        if deleted:
            log_event(
                "ADMIN_LOG_CLEANUP",
                deleted=deleted,
                retention_days=ADMIN_LOG_RETENTION_DAYS,
            )
        return int(deleted or 0)
    except Exception:
        db.session.rollback()
        return 0


def admin_required(permission=None, super_only=False):
    def _decorator(fn):
        @wraps(fn)
        def _wrapped(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()

            role = _effective_admin_role(current_user)
            if not role:
                abort(403)
            if super_only and role != "super_admin":
                abort(403)
            if permission and not _admin_has_permission(current_user, permission):
                abort(403)
            return fn(*args, **kwargs)
        return _wrapped
    return _decorator


def _ensure_user_created_at_column():
    try:
        inspector = inspect(db.engine)
        cols = {c["name"] for c in inspector.get_columns("user")}
        if "created_at" in cols:
            return False

        dialect = (db.engine.dialect.name or "").lower()
        if dialect in ("mysql", "mariadb"):
            db.session.execute(text("ALTER TABLE `user` ADD COLUMN created_at DATETIME NULL"))
            db.session.execute(text("UPDATE `user` SET created_at = UTC_TIMESTAMP() WHERE created_at IS NULL"))
        elif dialect == "sqlite":
            db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN created_at DATETIME"))
            db.session.execute(text("UPDATE \"user\" SET created_at = CURRENT_TIMESTAMP WHERE created_at IS NULL"))
        else:
            db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN created_at TIMESTAMP NULL"))
            db.session.execute(text("UPDATE \"user\" SET created_at = CURRENT_TIMESTAMP WHERE created_at IS NULL"))

        db.session.commit()
        print("Migracion aplicada: user.created_at agregado y rellenado para usuarios existentes.")
        return True
    except Exception as e:
        db.session.rollback()
        print(f"No se pudo aplicar migracion user.created_at: {e}")
        return False


# =========================================================
# VERIFICACIÓN DE CREACIÓN DE TABLAS
# =========================================================
def _ensure_user_is_active_account_column():
    try:
        inspector = inspect(db.engine)
        cols = {c["name"] for c in inspector.get_columns("user")}
        if "is_active_account" in cols:
            return False

        dialect = (db.engine.dialect.name or "").lower()
        if dialect in ("mysql", "mariadb"):
            db.session.execute(
                text("ALTER TABLE `user` ADD COLUMN is_active_account TINYINT(1) NOT NULL DEFAULT 1")
            )
            db.session.execute(
                text("UPDATE `user` SET is_active_account = 1 WHERE is_active_account IS NULL")
            )
        elif dialect == "sqlite":
            db.session.execute(
                text("ALTER TABLE \"user\" ADD COLUMN is_active_account BOOLEAN NOT NULL DEFAULT 1")
            )
            db.session.execute(
                text("UPDATE \"user\" SET is_active_account = 1 WHERE is_active_account IS NULL")
            )
        else:
            db.session.execute(
                text("ALTER TABLE \"user\" ADD COLUMN is_active_account BOOLEAN NOT NULL DEFAULT TRUE")
            )
            db.session.execute(
                text("UPDATE \"user\" SET is_active_account = TRUE WHERE is_active_account IS NULL")
            )

        db.session.commit()
        print("Migracion aplicada: user.is_active_account agregado y rellenado.")
        return True
    except Exception as e:
        db.session.rollback()
        print(f"No se pudo aplicar migracion user.is_active_account: {e}")
        return False


def _ensure_user_suspended_until_column():
    try:
        inspector = inspect(db.engine)
        cols = {c["name"] for c in inspector.get_columns("user")}
        if "suspended_until" in cols:
            return False

        dialect = (db.engine.dialect.name or "").lower()
        if dialect in ("mysql", "mariadb"):
            db.session.execute(text("ALTER TABLE `user` ADD COLUMN suspended_until DATETIME NULL"))
        elif dialect == "sqlite":
            db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN suspended_until DATETIME"))
        else:
            db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN suspended_until TIMESTAMP NULL"))

        db.session.commit()
        print("Migracion aplicada: user.suspended_until agregado.")
        return True
    except Exception as e:
        db.session.rollback()
        print(f"No se pudo aplicar migracion user.suspended_until: {e}")
        return False


print("=== CREANDO/MODIFICANDO TABLAS ===")
if ENVIRONMENT == "dev":
    print("Modelos detectados:", [model.__name__ for model in db.Model.__subclasses__()])

with app.app_context():
    db_init_retries = int(os.getenv("DB_INIT_RETRIES", "8"))
    db_init_delay_s = float(os.getenv("DB_INIT_RETRY_DELAY_S", "2.0"))

    for attempt in range(1, db_init_retries + 1):
        try:
            db.create_all()
            _ensure_user_created_at_column()
            _ensure_user_is_active_account_column()
            _ensure_user_suspended_until_column()
            _bootstrap_super_admin_roles()
            print("=== TABLAS CREADAS/VERIFICADAS ===")
            break
        except OperationalError as e:
            if attempt >= db_init_retries:
                raise
            print(
                f"⚠️ DB no disponible al iniciar (intento {attempt}/{db_init_retries}): {e}. "
                f"Reintentando en {db_init_delay_s}s..."
            )
            time.sleep(db_init_delay_s)
    
    # =========================================================
    # TEMPORAL: Limpiar datos de RateLimit con fechas mezcladas
    # (Ejecutar solo una vez, luego comentar o eliminar)
    # =========================================================
    #try:
     #   deleted = RateLimit.query.delete()
      #  db.session.commit()
       # if deleted > 0:
        #    print(f"🧹 Se limpiaron {deleted} registros antiguos de RateLimit")
    #except Exception as e:
     #   print(f"⚠️ Error al limpiar RateLimit: {e}")
    # =========================================================


# ✅ Dejar SOLO un user_loader
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))  # ✅ Cambio 4: Query.get() -> db.session.get()


# =========================================================
# 8) CONFIG IA (Gemini)
# =========================================================
instruccion_sistema = """
Eres Nexus, un asistente académico avanzado.
REGLAS:
1. Usa Markdown para todo (tablas, negritas, listas).
2. Si recibes una imagen, descríbela y ayuda con lo que contenga (matemáticas, texto, etc).
3. Sé amable y directo.
4. Entrega fórmulas limpias y legibles, sin símbolos basura ni escapes extraños.
"""
configuracion = {"temperature": 0.7}

# =========================================================
# Objetivo Fase 1 (solo esto ahorita)
# Quitar el error “Must supply api_key” y hacer que Gemini siempre analice la imagen.
# Paso A
# =========================================================
#model = genai.GenerativeModel(
 #   model_name='gemini-flash-latest',
  #  generation_config=configuracion,
   # system_instruction=instruccion_sistema
#)
#chat_session = model.start_chat(history=[])

# =========================================================
# 9) RESET LIMITS (anti-spam)
# =========================================================
RESET_COOLDOWN_SECONDS = 60
RESET_MAX_ATTEMPTS = 3
RESET_WINDOW_MINUTES = 30  # ventana en la que cuentan los 3 intentos

# =========================================================
# Fase 1 — Paso 4: Constantes de límites (fácil de ajustar)
# =========================================================
RESET_IP_MAX_ATTEMPTS = 8
RESET_IP_WINDOW_MINUTES = 30
RESET_IP_BLOCK_MINUTES = 15

LOGIN_MAX_ATTEMPTS = 7
LOGIN_WINDOW_MINUTES = 10
LOGIN_BLOCK_MINUTES = 10

# =========================================================
# Fase 2 — Paso 2.2: Constantes de Rate Limits
# =========================================================
# Login endpoint (anti-bots)
LOGIN_RL_MAX = 20              # 20 requests
LOGIN_RL_WINDOW_S = 60         # por 60s
LOGIN_RL_BLOCK_S = 60          # bloqueo 60s si abusa

# Forgot endpoint
FORGOT_RL_MAX = 10             # 10 requests
FORGOT_RL_WINDOW_S = 300       # por 5 min
FORGOT_RL_BLOCK_S = 300        # bloqueo 5 min

# Chat endpoint (protege Gemini)
CHAT_RL_MAX = 12               # 12 mensajes
CHAT_RL_WINDOW_S = 60          # por 60s
CHAT_RL_BLOCK_S = 60           # bloqueo 60s

# Límites de payload (seguridad + costos)
CHAT_MAX_TEXT_CHARS = 2000
CHAT_MAX_IMAGE_BYTES = 8 * 1024 * 1024   # 8MB
AI_REQUEST_TIMEOUT_S = int((os.getenv("AI_REQUEST_TIMEOUT_S") or "15").strip())
AI_MAX_KEY_RETRIES = int((os.getenv("AI_MAX_KEY_RETRIES") or "1").strip())
AI_MODEL_CANDIDATES = [
    x.strip()
    for x in (os.getenv("AI_MODEL_CANDIDATES") or "gemini-2.5-flash,gemini-flash-lite-latest,gemini-flash-latest").split(",")
    if x.strip()
]
RESOURCE_HTTP_TIMEOUT_S = int((os.getenv("RESOURCE_HTTP_TIMEOUT_S") or "6").strip())
WIKI_ENABLED = (os.getenv("WIKI_ENABLED") or "1").strip() == "1"
WIKI_LANG = (os.getenv("WIKI_LANG") or "es").strip().lower()
try:
    WIKI_HINT_PROB = float((os.getenv("WIKI_HINT_PROB") or "0.45").strip())
except Exception:
    WIKI_HINT_PROB = 0.45
WIKI_HINT_PROB = max(0.0, min(1.0, WIKI_HINT_PROB))
YOUTUBE_API_KEY = (os.getenv("YOUTUBE_API_KEY") or "").strip()
YOUTUBE_ENABLED = ((os.getenv("YOUTUBE_ENABLED") or "1").strip() == "1") and bool(YOUTUBE_API_KEY)
try:
    YOUTUBE_INCLUDE_PROB = float((os.getenv("YOUTUBE_INCLUDE_PROB") or "0.35").strip())
except Exception:
    YOUTUBE_INCLUDE_PROB = 0.35
YOUTUBE_INCLUDE_PROB = max(0.0, min(1.0, YOUTUBE_INCLUDE_PROB))
YOUTUBE_MAX_RESULTS = max(1, min(3, int((os.getenv("YOUTUBE_MAX_RESULTS") or "2").strip())))


# =========================================================
# 10) HTML del correo (Reset)
# =========================================================
def build_reset_email_html(name: str, link: str) -> str:
    return f"""
<div style="margin:0; padding:0; background:#0b1220; font-family:Segoe UI, Arial, sans-serif;">
  <div style="max-width:640px; margin:0 auto; padding:28px 16px;">

    <div style="
      background:linear-gradient(180deg,#0f172a 0%, #0b1220 100%);
      border:1px solid #1f2a44;
      border-radius:18px;
      overflow:hidden;
      box-shadow:0 12px 30px rgba(0,0,0,.35);
    ">

      <div style="padding:22px 20px; text-align:center; border-bottom:1px solid #1f2a44;">
        <div style="font-size:30px; font-weight:900; letter-spacing:1px; color:#22d3ee;">NEXUS</div>
        <div style="margin-top:6px; color:#94a3b8; font-size:13px;">Recuperación de contraseña</div>

        <div style="margin-top:14px;">
          <span style="
            display:inline-block;
            padding:9px 14px;
            border-radius:999px;
            background:rgba(34,211,238,.08);
            border:1px solid rgba(34,211,238,.35);
            color:#cbd5e1;
            font-size:12px;
            font-weight:700;
          ">
            Enlace válido por <span style="color:#7dd3fc; font-weight:900;">20 minutos</span>
          </span>
        </div>
      </div>

      <div style="padding:22px 20px; color:#e2e8f0;">
        <p style="margin:0 0 12px 0; font-size:15px;">
          Hola <b style="color:#ffffff;">{name}</b>,
        </p>

        <p style="margin:0 0 16px 0; font-size:14px; color:#cbd5e1; line-height:1.65;">
          Recibimos una solicitud para restablecer tu contraseña. Si fuiste tú, presiona el botón:
        </p>

        <div style="text-align:center; margin:18px 0 14px 0;">
          <a href="{link}" style="
            display:inline-block;
            padding:13px 18px;
            border-radius:12px;
            background:#22d3ee;
            color:#06212a;
            text-decoration:none;
            font-weight:900;
            font-size:14px;
            box-shadow:0 10px 22px rgba(34,211,238,.20);
          ">
            Restablecer contraseña
          </a>
        </div>

        <p style="margin:0; font-size:12.5px; color:#94a3b8; line-height:1.6;">
          Si tú no hiciste esta solicitud, puedes ignorar este correo.
        </p>

        <div style="margin-top:18px; padding-top:16px; border-top:1px solid #1f2a44;">
          <div style="font-size:12px; color:#94a3b8; margin-bottom:10px;">
            Si el botón no funciona, copia y pega este enlace:
          </div>

          <div style="
            word-break:break-all;
            padding:12px 12px;
            border-radius:12px;
            background:#07101f;
            border:1px solid #1f2a44;
            color:#cbd5e1;
            font-size:12.5px;
            line-height:1.6;
          ">{link}</div>
        </div>

        <div style="margin-top:18px; padding-top:16px; border-top:1px solid #1f2a44;">
          <div style="
            display:inline-block;
            font-size:11px;
            letter-spacing:.6px;
            text-transform:uppercase;
            color:#94a3b8;
            background:#07101f;
            border:1px solid #1f2a44;
            padding:6px 10px;
            border-radius:999px;
          ">
            Soporte Nexus
          </div>

          <p style="margin:12px 0 12px 0; font-size:12.5px; color:#94a3b8; line-height:1.6;">
            Este es un correo automático. Si necesitas ayuda, contáctanos:
          </p>

          <div style="text-align:center; margin:8px 0 2px 0;">
            <a href="mailto:jonathandavidloza@gmail.com" style="
              display:inline-block;
              margin:6px 6px;
              padding:10px 14px;
              border-radius:12px;
              background:#07101f;
              border:1px solid #1f2a44;
              color:#e2e8f0;
              text-decoration:none;
              font-weight:800;
              font-size:13px;
            ">Escribir a soporte</a>

            <a href="https://wa.me/50364254348?text=Hola%20Nexus%2C%20necesito%20ayuda%20con%20mi%20cuenta." style="
              display:inline-block;
              margin:6px 6px;
              padding:10px 14px;
              border-radius:12px;
              background:#22d3ee;
              border:1px solid rgba(34,211,238,.55);
              color:#06212a;
              text-decoration:none;
              font-weight:900;
              font-size:13px;
              box-shadow:0 10px 22px rgba(34,211,238,.18);
            ">WhatsApp soporte</a>
          </div>

          <div style="text-align:center; color:#64748b; font-size:12px; margin-top:14px;">
            © 2026 Nexus • Seguridad de cuenta
          </div>
        </div>

      </div>
    </div>
  </div>
</div>
"""


# =========================================================
# 11) ENVÍO DEL CORREO (dev / smtp / brevo_api)
# =========================================================
def send_reset_link(email, name, link):
    mode = (RESET_MODE or "dev").strip().lower()

    if mode == "dev":
        print("\n==============================")
        print("🔗 LINK RESET (DEV):", link)
        print("==============================\n")
        log_event("EMAIL_SENT", provider="dev_console", to=email, ok=True)
        return True

    if mode == "brevo_api":
        try:
            subject = "Recuperación de contraseña - Nexus"
            text_body = f"""Hola {name},

Recibimos una solicitud para restablecer tu contraseña.
Este enlace es válido por 20 minutos:

{link}

Si tú no hiciste esta solicitud, ignora este mensaje.

---
Soporte:
Correo: jonathandavidloza@gmail.com
WhatsApp: https://wa.me/50364254348
"""
            html_body = build_reset_email_html(name=name, link=link)

            url = "https://api.brevo.com/v3/smtp/email"
            headers = {
                "accept": "application/json",
                "api-key": BREVO_API_KEY,
                "content-type": "application/json",
            }
            payload = {
                "sender": {"name": BREVO_SENDER_NAME, "email": BREVO_SENDER_EMAIL},
                "to": [{"email": email, "name": name}],
                "subject": subject,
                "htmlContent": html_body,
                "textContent": text_body,
                "replyTo": {"email": "jonathandavidloza@gmail.com", "name": "Soporte Nexus"},
            }

            r = requests.post(url, headers=headers, json=payload, timeout=10)

            if 200 <= r.status_code < 300:
                log_event("EMAIL_SENT", provider="brevo_api", to=email, ok=True, status=r.status_code)
                return True

            print("❌ Error Brevo API:", r.status_code, r.text)
            print("🔗 LINK RESET (FALLBACK):", link)
            log_event("EMAIL_SENT", provider="brevo_api", to=email, ok=False, status=r.status_code)
            return False

        except Exception as e:
            print("❌ Exception Brevo API:", repr(e))
            print("🔗 LINK RESET (FALLBACK):", link)
            log_event("EMAIL_SENT", provider="brevo_api", to=email, ok=False, error=str(e))
            return False

    try:
        msg = MailMessage(
            subject="Recuperación de contraseña - Nexus",
            recipients=[email]
        )

        msg.reply_to = "jonathandavidloza@gmail.com"

        msg.body = f"""Hola {name},

Recibimos una solicitud para restablecer tu contraseña.
Este enlace es válido por 20 minutos:

{link}

Si tú no hiciste esta solicitud, ignora este mensaje.
"""

        msg.html = build_reset_email_html(name=name, link=link)

        mail.send(msg)
        log_event("EMAIL_SENT", provider="smtp", to=email, ok=True)
        return True

    except Exception as e:
        print("❌ Error enviando correo (SMTP/Brevo):", repr(e))
        print("🔗 LINK RESET (FALLBACK DEV):", link)
        log_event("EMAIL_SENT", provider="smtp", to=email, ok=False, error=str(e))
        return True


# =========================================================
# 12) CONTROL DE INTENTOS (anti-spam) - CORREGIDO CON NAIVE UTC
# =========================================================
def can_send_reset(email: str):
    now = utcnow_naive()  # ✅ Cambiado a naive UTC
    rr = ResetRequest.query.filter_by(email=email).first()

    if not rr:
        rr = ResetRequest(email=email, last_sent_at=None, attempts=0, first_attempt_at=None)
        db.session.add(rr)
        db.session.commit()

    first = to_naive_utc(rr.first_attempt_at)
    last = to_naive_utc(rr.last_sent_at)

    if first and now - first > timedelta(minutes=RESET_WINDOW_MINUTES):
        rr.attempts = 0
        rr.first_attempt_at = None
        rr.last_sent_at = None
        db.session.commit()

    if last and (now - last).total_seconds() < RESET_COOLDOWN_SECONDS:
        wait = RESET_COOLDOWN_SECONDS - int((now - last).total_seconds())
        return False, wait, (rr.attempts >= RESET_MAX_ATTEMPTS)

    if rr.attempts >= RESET_MAX_ATTEMPTS:
        return False, 0, True

    return True, 0, False


def register_reset_sent(email: str):
    now = utcnow_naive()  # ✅ Cambiado a naive UTC
    rr = ResetRequest.query.filter_by(email=email).first()
    if not rr:
        rr = ResetRequest(email=email)
        db.session.add(rr)

    if rr.attempts == 0 or rr.first_attempt_at is None:
        rr.first_attempt_at = now

    rr.attempts = (rr.attempts or 0) + 1
    rr.last_sent_at = now
    db.session.commit()


def can_send_reset_ip(ip: str):
    now = utcnow_naive()  # ✅ Cambiado a naive UTC
    manual_ip_block = _active_security_block("ip", ip, now)
    if manual_ip_block:
        wait = _security_block_wait_seconds(manual_ip_block, now)
        return False, wait, True

    row = ResetIPRequest.query.filter_by(ip=ip).first()
    if not row:
        row = ResetIPRequest(ip=ip)
        db.session.add(row)
        db.session.commit()

    blocked_until = to_naive_utc(row.blocked_until)
    first_attempt_at = to_naive_utc(row.first_attempt_at)
    last_sent_at = to_naive_utc(row.last_sent_at)

    if blocked_until and now < blocked_until:
        wait = int((blocked_until - now).total_seconds())
        return False, wait, True

    if first_attempt_at and now - first_attempt_at > timedelta(minutes=RESET_IP_WINDOW_MINUTES):
        row.attempts = 0
        row.first_attempt_at = None
        row.last_sent_at = None
        row.blocked_until = None
        db.session.commit()

    if (row.attempts or 0) >= RESET_IP_MAX_ATTEMPTS:
        row.blocked_until = now + timedelta(minutes=RESET_IP_BLOCK_MINUTES)
        db.session.commit()
        wait = int((row.blocked_until - now).total_seconds())
        return False, wait, True

    return True, 0, False


def register_reset_ip_sent(ip: str):
    now = utcnow_naive()  # ✅ Cambiado a naive UTC
    row = ResetIPRequest.query.filter_by(ip=ip).first()
    if not row:
        row = ResetIPRequest(ip=ip)
        db.session.add(row)

    if (row.attempts or 0) == 0 or row.first_attempt_at is None:
        row.first_attempt_at = now

    row.attempts = (row.attempts or 0) + 1
    row.last_sent_at = now
    db.session.commit()


def can_login(ip: str, email: str):
    now = utcnow_naive()  # ✅ Cambiado a naive UTC
    manual_email_block = _active_security_block("email", email, now)
    manual_ip_block = _active_security_block("ip", ip, now)
    if manual_email_block or manual_ip_block:
        wait = max(
            _security_block_wait_seconds(manual_email_block, now),
            _security_block_wait_seconds(manual_ip_block, now),
        )
        return False, wait

    row = LoginAttempt.query.filter_by(ip=ip, email=email).first()
    if not row:
        row = LoginAttempt(ip=ip, email=email)
        db.session.add(row)
        db.session.commit()

    blocked_until = to_naive_utc(row.blocked_until)
    first_attempt_at = to_naive_utc(row.first_attempt_at)

    if blocked_until and now < blocked_until:
        wait = int((blocked_until - now).total_seconds())
        return False, wait

    if first_attempt_at and now - first_attempt_at > timedelta(minutes=LOGIN_WINDOW_MINUTES):
        row.attempts = 0
        row.first_attempt_at = None
        row.blocked_until = None
        db.session.commit()

    if (row.attempts or 0) >= LOGIN_MAX_ATTEMPTS:
        row.blocked_until = now + timedelta(minutes=LOGIN_BLOCK_MINUTES)
        db.session.commit()
        wait = int((row.blocked_until - now).total_seconds())
        return False, wait

    return True, 0


def register_login_fail(ip: str, email: str):
    now = utcnow_naive()  # ✅ Cambiado a naive UTC
    row = LoginAttempt.query.filter_by(ip=ip, email=email).first()
    if not row:
        row = LoginAttempt(ip=ip, email=email)
        db.session.add(row)

    if (row.attempts or 0) == 0 or row.first_attempt_at is None:
        row.first_attempt_at = now

    row.attempts = (row.attempts or 0) + 1
    db.session.commit()


def clear_login_attempts(ip: str, email: str):
    row = LoginAttempt.query.filter_by(ip=ip, email=email).first()
    if row:
        db.session.delete(row)
        db.session.commit()


# =========================================================
# Fase 2 — Paso 2.3: Helpers de Rate Limit (DB persistente)
# =========================================================
def _rl_key(endpoint: str, ip: str, user_id: int | None = None) -> str:
    """
    Genera una clave única para rate limiting.
    Formato: endpoint:ip:user_id (si hay usuario)
    """
    uid = user_id or 0
    return f"{endpoint}:{ip}:{uid}"


def _as_utc_naive(dt):
    """
    Convierte cualquier datetime (aware o naive) a naive en UTC.
    - Si ya es naive, se asume que es UTC.
    """
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt  # asumimos UTC
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


def rate_limit_check(key: str, max_count: int, window_seconds: int, block_seconds: int):
    """
    Rate limit persistente en DB.
    Retorna: (ok, wait_s)
    """
    # Trabajamos EN NAIVE UTC para evitar choque con valores DB (naive)
    # FIX: evitar datetime.utcnow() (deprecated)
    now = utcnow_naive()

    row = RateLimit.query.filter_by(key=key).first()
    if not row:
        row = RateLimit(key=key, window_start=now, count=0, blocked_until=None)
        db.session.add(row)
        db.session.commit()

    window_start = _as_utc_naive(row.window_start)
    blocked_until = _as_utc_naive(row.blocked_until)

    # si está bloqueado
    if blocked_until and now < blocked_until:
        wait = int((blocked_until - now).total_seconds())
        return False, wait

    # reset de ventana si pasó el tiempo
    if window_start and (now - window_start).total_seconds() > window_seconds:
        row.window_start = now
        row.count = 0
        row.blocked_until = None
        db.session.commit()
        return True, 0

    # contar el request actual
    row.count = (row.count or 0) + 1

    if row.count > max_count:
        row.blocked_until = now + timedelta(seconds=block_seconds)
        db.session.commit()
        wait = int((row.blocked_until - now).total_seconds())
        return False, wait

    db.session.commit()
    return True, 0


def _set_login_help_mode(mode: str = "", email_hint: str = ""):
    """
    mode:
      - "forgot": mostrar solo enlace de recuperacion
      - "support": mostrar solo enlace de soporte
      - otro/empty: ocultar ambos
    """
    mode = (mode or "").strip().lower()
    email_hint = (email_hint or "").strip().lower()

    if mode == "forgot":
        session["show_forgot"] = True
        session["show_inactive_support"] = False
        session.pop("support_email_hint", None)
        return

    if mode == "support":
        session["show_forgot"] = False
        session["show_inactive_support"] = True
        if email_hint:
            session["support_email_hint"] = email_hint
        return

    session.pop("show_forgot", None)
    session.pop("show_inactive_support", None)
    session.pop("support_email_hint", None)


def _build_support_whatsapp_link():
    if not SUPPORT_WHATSAPP:
        return ""
    base = f"https://wa.me/{SUPPORT_WHATSAPP}"
    email_hint = (session.get("support_email_hint") or "").strip()
    msg = "Hola, necesito ayuda con mi cuenta de Nexus."
    if email_hint:
        msg += f" Mi correo es: {email_hint}."
    msg += " Me aparece como desactivada."
    return f"{base}?{urlencode({'text': msg})}"


# =========================================================
# 13) RUTAS AUTH (Login/Register/Reset)
# =========================================================
@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for('home'))

    show_forgot = session.get("show_forgot", False)
    show_inactive_support = session.get("show_inactive_support", False)
    if show_inactive_support:
        show_forgot = False
    support_whatsapp_link = _build_support_whatsapp_link()

    if request.method == 'POST':
        email = (request.form.get('email') or "").strip().lower()
        password = request.form.get('password') or ""

        ip = get_client_ip()

        # =========================================================
        # Fase 2 - Paso 2.4: Rate limit para /login (anti-bots)
        # =========================================================
        rl_ok, rl_wait = rate_limit_check(
            key=_rl_key("login", ip, None),
            max_count=LOGIN_RL_MAX,
            window_seconds=LOGIN_RL_WINDOW_S,
            block_seconds=LOGIN_RL_BLOCK_S
        )
        if not rl_ok:
            flash(f"Demasiadas solicitudes. Espera {rl_wait} segundos e intenta de nuevo.", "error")
            _set_login_help_mode("forgot")
            return render_template(
                'login.html',
                show_forgot=True,
                show_inactive_support=False,
                support_whatsapp_link="",
            )
        # =========================================================

        ok_login, wait_login = can_login(ip, email)
        if not ok_login:
            manual_email_block = _active_security_block("email", email)
            manual_ip_block = _active_security_block("ip", ip)
            if manual_email_block or manual_ip_block:
                log_event("LOGIN_BLOCKED", email=email, ip=ip, reason="manual_security_block", wait_s=wait_login)
                flash(f"Acceso bloqueado por seguridad. Espera {wait_login} segundos o contacta soporte.", "error")
                _set_login_help_mode("support", email_hint=email)
            else:
                log_event("LOGIN_BLOCKED", email=email, ip=ip, wait_s=wait_login)
                flash(f"Demasiados intentos. Espera {wait_login} segundos o usa recuperacion de contrasena.", "error")
                _set_login_help_mode("forgot")
            return render_template(
                'login.html',
                show_forgot=session.get("show_forgot", False),
                show_inactive_support=session.get("show_inactive_support", False),
                support_whatsapp_link=_build_support_whatsapp_link(),
            )

        user = User.query.filter_by(email=email).first()

        if not user:
            log_event("LOGIN_FAIL", email=email, ip=ip, reason="no_user")
            flash('Este correo no esta registrado.', 'error')
            _set_login_help_mode("forgot")
            register_login_fail(ip, email)

        elif not check_password_hash(user.password, password):
            log_event("LOGIN_FAIL", email=email, ip=ip, reason="bad_password")
            flash('Contrasena incorrecta. Intentalo de nuevo.', 'error')
            _set_login_help_mode("forgot")
            register_login_fail(ip, email)

        elif not bool(user.is_active_account):
            log_event("LOGIN_BLOCKED", email=email, ip=ip, reason="inactive_account", user_id=user.id)
            flash('Tu cuenta esta desactivada. Contacta al administrador.', 'error')
            _set_login_help_mode("support", email_hint=email)

        elif _active_suspension_until(user):
            until = _active_suspension_until(user)
            log_event(
                "LOGIN_BLOCKED",
                email=email,
                ip=ip,
                reason="suspended_account",
                user_id=user.id,
                suspended_until=_format_dt_human(until),
            )
            flash(
                f"Tu cuenta esta suspendida hasta {_format_dt_human(until)}. Contacta al administrador.",
                'error'
            )
            _set_login_help_mode("support", email_hint=email)

        else:
            log_event("LOGIN_OK", email=email, ip=ip, user_id=user.id)
            _set_login_help_mode("")
            clear_login_attempts(ip, email)
            login_user(user)
            session["login_at_ts"] = int(time.time())
            _ensure_super_admin_membership(user)
            return redirect(url_for('home'))

        show_forgot = session.get("show_forgot", False)
        show_inactive_support = session.get("show_inactive_support", False)
        if show_inactive_support:
            show_forgot = False
        support_whatsapp_link = _build_support_whatsapp_link()

    return render_template(
        'login.html',
        show_forgot=show_forgot,
        show_inactive_support=show_inactive_support,
        support_whatsapp_link=support_whatsapp_link,
    )


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'GET':
        return redirect(url_for('login_page', tab='register'))

    email = request.form.get('email')
    name = request.form.get('nombre')
    password = request.form.get('password')

    if User.query.filter_by(email=email).first():
        flash('Correo en uso. Este correo ya está registrado, usa otro por favor.', 'error')
        return redirect(url_for('login_page', tab='register'))

    if not EMAIL_RE.match(email or ""):
        flash('Correo inválido. Usa un formato como nombre@dominio.com.', 'error')
        return redirect(url_for('login_page', tab='register'))

    if not PASSWORD_RE.match(password or ""):
        flash('La contraseña debe tener al menos 6 caracteres e incluir un símbolo.', 'error')
        return redirect(url_for('login_page', tab='register'))

    new_user = User(email=email, name=name, password=generate_password_hash(password, method='scrypt'))
    db.session.add(new_user)
    db.session.commit()

    login_user(new_user)
    session["login_at_ts"] = int(time.time())
    _ensure_super_admin_membership(new_user)
    return redirect(url_for('home'))


@app.route('/logout')
@login_required
def logout():
    logout_user()
    # Mantener datos de enlaces compartidos activos en esta sesión del navegador.
    # Solo limpiamos claves de autenticación principal.
    for key in ["_user_id", "_fresh", "_id", "remember_token", "login_at_ts"]:
        session.pop(key, None)
    return redirect(url_for('login_page'))


@app.route('/forgot', methods=['GET', 'POST'])
def forgot_password():
    show_support = False

    if request.method == 'POST':
        email = (request.form.get('email') or "").strip().lower()
        user = User.query.filter_by(email=email).first()

        log_event("RESET_REQUEST", email=email, ip=get_client_ip())

        ok, wait, support = can_send_reset(email)
        show_support = support

        ip = get_client_ip()

        # =========================================================
        # Fase 2 — Paso 2.5: Rate limit para /forgot (anti-spam)
        # =========================================================
        rl_ok, rl_wait = rate_limit_check(
            key=_rl_key("forgot", ip, None),
            max_count=FORGOT_RL_MAX,
            window_seconds=FORGOT_RL_WINDOW_S,
            block_seconds=FORGOT_RL_BLOCK_S
        )
        if not rl_ok:
            show_support = True
            flash(f"Demasiadas solicitudes. Espera {rl_wait} segundos o contacta soporte.", "error")
            return render_template("forgot_password.html", show_support=show_support)
        # =========================================================

        ok_ip, wait_ip, blocked_ip = can_send_reset_ip(ip)
        if not ok_ip:
            log_event("RESET_BLOCKED", email=email, ip=ip, reason="ip_limit", wait_s=wait_ip)
            show_support = True
            flash("Demasiadas solicitudes desde tu red. Intenta más tarde o contacta soporte técnico.", "error")
            return render_template("forgot_password.html", show_support=show_support)

        if not ok:
            reason = "cooldown" if wait > 0 else "max_attempts"
            log_event("RESET_BLOCKED", email=email, ip=ip, reason=reason, wait_s=wait)

            if wait > 0:
                flash(f"Espera {wait} segundos para volver a enviar el enlace.", "error")
            else:
                flash("Se alcanzó el máximo de intentos. Contacta soporte técnico.", "error")
            return render_template("forgot_password.html", show_support=show_support)

        if user:
            token = serializer.dumps(email, salt="reset-password")
            link = url_for('reset_password', token=token, _external=True)

            log_event("RESET_SEND_ATTEMPT", email=email, ip=ip, user_id=user.id)
            sent = send_reset_link(email=user.email, name=user.name, link=link)
            if sent:
                log_event("RESET_SENT", email=email, ip=ip, user_id=user.id)
                register_reset_sent(email)
                register_reset_ip_sent(ip)
            else:
                log_event("RESET_SEND_FAIL", email=email, ip=ip, user_id=user.id)

        flash("Si el correo existe, te enviamos un enlace para recuperar tu contraseña. Revisa bandeja y spam.", "success")
        return render_template("forgot_password.html", show_support=show_support)

    return render_template('forgot_password.html', show_support=show_support)


@app.context_processor
def inject_admin_nav():
    if not current_user.is_authenticated:
        return {
            "can_access_admin": False,
            "admin_role_label": None,
            "is_super_admin": False,
        }

    role = _effective_admin_role(current_user)
    return {
        "can_access_admin": bool(role),
        "admin_role_label": role,
        "is_super_admin": role == "super_admin",
    }


@app.route('/reset/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        email = serializer.loads(token, salt="reset-password", max_age=RESET_TOKEN_MAX_AGE)
    except SignatureExpired:
        flash("Este enlace ya expiró. Solicita uno nuevo para restablecer tu contraseña.", "error")
        return redirect(url_for('forgot_password'))
    except BadSignature:
        flash("Enlace inválido.", "error")
        return redirect(url_for('forgot_password'))

    user = User.query.filter_by(email=email).first()
    if not user:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for('login_page'))

    if request.method == 'POST':
        new_password = request.form.get('password') or ""

        if not PASSWORD_RE.match(new_password):
            flash("La contraseña debe tener al menos 6 caracteres e incluir un símbolo.", "error")
            return redirect(url_for('reset_password', token=token))

        user.password = generate_password_hash(new_password, method='scrypt')
        db.session.commit()

        _set_login_help_mode("")

        flash("Contraseña actualizada. Ya puedes iniciar sesión.", "success")
        return redirect(url_for('login_page'))

    return render_template('reset_password.html', token=token)


# =========================================================
# 14) CHAT
# =========================================================
@app.route('/')
@app.route('/c/<int:chat_id>')
@login_required
def home(chat_id=None):
    mis_conversaciones = Conversation.query.filter_by(user_id=current_user.id).order_by(Conversation.created_at.desc()).all()
    mensajes_actuales = []
    chat_activo = None

    if chat_id:
        # FIX: SQLAlchemy 2.0 reemplaza Query.get() por session.get()
        chat_activo = db.session.get(Conversation, chat_id)
        if chat_activo and chat_activo.user_id == current_user.id:
            mensajes_actuales = Message.query.filter_by(conversation_id=chat_id).order_by(Message.timestamp).all()
        else:
            return redirect(url_for('home'))

    return render_template(
        'index.html',
        name=current_user.name,
        email=current_user.email,
        conversations=mis_conversaciones,
        chat_history=mensajes_actuales,
        active_chat=chat_activo
    )


def _admin_stats():
    return {
        "total_users": User.query.count(),
        "total_chats": Conversation.query.count(),
        "total_messages": Message.query.count(),
        "total_admins": AdminRole.query.filter_by(is_active=True).count(),
    }


def _admin_dashboard_charts(days=7):
    now_utc = utcnow_naive()
    start_day = (now_utc - timedelta(days=max(1, days) - 1)).replace(hour=0, minute=0, second=0, microsecond=0)

    day_labels = []
    day_map = {}
    for i in range(max(1, days)):
        d = (start_day + timedelta(days=i)).date()
        key = d.strftime("%Y-%m-%d")
        day_labels.append(key)
        day_map[key] = 0

    rows = (
        Message.query
        .filter(Message.timestamp >= start_day)
        .order_by(Message.timestamp.asc())
        .all()
    )
    for m in rows:
        if not m.timestamp:
            continue
        key = m.timestamp.date().strftime("%Y-%m-%d")
        if key in day_map:
            day_map[key] += 1
    message_day_values = [day_map[k] for k in day_labels]

    active_users = User.query.filter(User.is_active_account == True).count()  # noqa: E712
    inactive_users = User.query.filter(User.is_active_account == False).count()  # noqa: E712

    ip_rows = (
        LoginAttempt.query
        .filter(LoginAttempt.attempts > 0)
        .order_by(LoginAttempt.attempts.desc(), LoginAttempt.id.desc())
        .limit(8)
        .all()
    )
    ip_labels = [r.ip or "-" for r in ip_rows]
    ip_values = [int(r.attempts or 0) for r in ip_rows]

    return {
        "message_day_labels": day_labels,
        "message_day_values": message_day_values,
        "account_status_labels": ["Activas", "Desactivadas"],
        "account_status_values": [active_users, inactive_users],
        "failed_ip_labels": ip_labels,
        "failed_ip_values": ip_values,
    }


def _format_uptime_compact(start_dt, now_dt):
    if not start_dt or not now_dt or now_dt < start_dt:
        return "-"
    total_seconds = int((now_dt - start_dt).total_seconds())
    days, rem = divmod(total_seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, _ = divmod(rem, 60)
    parts = []
    if days:
        parts.append(f"{days}d")
    if hours or days:
        parts.append(f"{hours}h")
    parts.append(f"{minutes}m")
    return " ".join(parts)


def _admin_system_health():
    now_utc = utcnow_naive()
    items = []

    # Base de datos + latencia simple de ping
    db_status = "off"
    db_label = "Error"
    db_detail = "Sin respuesta"
    try:
        t0 = time.perf_counter()
        db.session.execute(text("SELECT 1"))
        latency_ms = int((time.perf_counter() - t0) * 1000)
        if latency_ms <= 120:
            db_status, db_label = "ok", "Conectada"
        elif latency_ms <= 350:
            db_status, db_label = "warn", "Lenta"
        else:
            db_status, db_label = "warn", "Latencia alta"
        db_detail = f"Ping: {latency_ms} ms"
    except Exception:
        pass
    items.append({
        "name": "Base de datos",
        "status": db_status,
        "label": db_label,
        "detail": db_detail,
        "icon": "fa-database",
    })

    # Gemini
    gemini_count = len([k for k in LISTA_DE_CLAVES if k])
    if gemini_count >= 3:
        gem_status, gem_label = "ok", "Lista"
    elif gemini_count >= 1:
        gem_status, gem_label = "warn", "Parcial"
    else:
        gem_status, gem_label = "off", "Sin claves"
    items.append({
        "name": "Gemini",
        "status": gem_status,
        "label": gem_label,
        "detail": f"Claves cargadas: {gemini_count}",
        "icon": "fa-robot",
    })

    # Correo
    mode = (RESET_MODE or "dev").strip().lower()
    if mode == "brevo_api":
        mail_ok = bool(BREVO_API_KEY and BREVO_SENDER_EMAIL)
        mail_status, mail_label = ("ok", "Brevo API") if mail_ok else ("off", "Brevo incompleto")
    elif mode == "smtp":
        mail_ok = bool(MAIL_SERVER and MAIL_USERNAME and MAIL_PASSWORD)
        mail_status, mail_label = ("ok", "SMTP listo") if mail_ok else ("off", "SMTP incompleto")
    else:
        mail_ok = True
        mail_status, mail_label = "warn", "Modo dev"
    items.append({
        "name": "Correo",
        "status": mail_status,
        "label": mail_label,
        "detail": f"Modo: {mode}",
        "icon": "fa-envelope",
    })

    # Cloudinary
    cfg = cloudinary.config()
    cloud_ok = bool(cfg and cfg.cloud_name and (cfg.api_key or CLOUDINARY_URL))
    items.append({
        "name": "Cloudinary",
        "status": "ok" if cloud_ok else "off",
        "label": "Listo" if cloud_ok else "No configurado",
        "detail": f"Cloud: {cfg.cloud_name if cfg and cfg.cloud_name else '-'}",
        "icon": "fa-cloud",
    })

    # Uptime del proceso actual
    items.append({
        "name": "Uptime",
        "status": "ok",
        "label": _format_uptime_compact(APP_STARTED_AT, now_utc),
        "detail": f"Inicio: {_format_dt_human(APP_STARTED_AT)} UTC",
        "icon": "fa-clock",
    })

    # Seguridad reciente (señales rápidas)
    try:
        high_risk = LoginAttempt.query.filter(LoginAttempt.attempts >= 5).count()
    except Exception:
        high_risk = 0
    if high_risk >= 20:
        sec_status, sec_label = "off", "Alta actividad"
    elif high_risk >= 5:
        sec_status, sec_label = "warn", "Moderada"
    else:
        sec_status, sec_label = "ok", "Normal"
    items.append({
        "name": "Seguridad reciente",
        "status": sec_status,
        "label": sec_label,
        "detail": f"IPs con >=5 intentos: {high_risk}",
        "icon": "fa-shield-halved",
    })

    return items


def _admin_admins_data():
    admins_rows = (
        db.session.query(AdminRole, User)
        .join(User, User.id == AdminRole.user_id)
        .filter(AdminRole.is_active == True)  # noqa: E712
        .order_by(AdminRole.created_at.desc())
        .all()
    )
    rows = []
    for role_row, user_row in admins_rows:
        perm_codes = sorted(_loads_permissions(role_row.permissions_json))
        rows.append({
            "role_row": role_row,
            "user_row": user_row,
            "permission_codes": perm_codes,
            "permission_labels": [_permission_label_es(p) for p in perm_codes],
        })
    return rows


def _admin_users_data():
    chats_per_user_sq = (
        db.session.query(
            Conversation.user_id.label("uid"),
            func.count(Conversation.id).label("chat_count"),
        )
        .group_by(Conversation.user_id)
        .subquery()
    )

    messages_per_user_sq = (
        db.session.query(
            Conversation.user_id.label("uid"),
            func.count(Message.id).label("message_count"),
        )
        .outerjoin(Message, Message.conversation_id == Conversation.id)
        .group_by(Conversation.user_id)
        .subquery()
    )

    all_users = (
        db.session.query(
            User.id,
            User.name,
            User.email,
            User.created_at,
            User.is_active_account,
            User.suspended_until,
            func.coalesce(chats_per_user_sq.c.chat_count, 0).label("chat_count"),
            func.coalesce(messages_per_user_sq.c.message_count, 0).label("message_count"),
        )
        .outerjoin(chats_per_user_sq, chats_per_user_sq.c.uid == User.id)
        .outerjoin(messages_per_user_sq, messages_per_user_sq.c.uid == User.id)
        .order_by(User.created_at.desc())
        .all()
    )
    return all_users


def _admin_recent_logs(limit=30):
    return (
        db.session.query(AdminAuditLog, User)
        .outerjoin(User, User.id == AdminAuditLog.actor_user_id)
        .order_by(AdminAuditLog.created_at.desc())
        .limit(limit)
        .all()
    )


def _mask_email_for_logs(email: str):
    email = (email or "").strip()
    if "@" not in email:
        return email
    local, domain = email.split("@", 1)
    if len(local) <= 2:
        local_masked = local[:1] + "*"
    else:
        local_masked = local[:2] + ("*" * max(2, len(local) - 2))
    domain_parts = domain.split(".")
    if domain_parts and domain_parts[0]:
        dom0 = domain_parts[0]
        domain_parts[0] = dom0[:1] + ("*" * max(2, len(dom0) - 1))
    return local_masked + "@" + ".".join(domain_parts)


def _mask_sensitive_text(text: str):
    raw = (text or "").strip()
    if not raw:
        return "-"
    # Claves/tokens/passwords declarados en pares key=value
    masked = re.sub(
        r"(?i)\b(api[_-]?key|token|secret|password|authorization)\s*=\s*([^;,\s]+)",
        r"\1=***",
        raw,
    )
    # Bearer tokens
    masked = re.sub(r"(?i)\bBearer\s+[A-Za-z0-9\-._~+/]+=*", "Bearer ***", masked)
    # Gemini/Google API keys tipo AIza...
    masked = re.sub(r"AIza[0-9A-Za-z\-_]{20,}", "AIza***", masked)
    # Enmascarar correos
    masked = re.sub(
        r"([A-Za-z0-9._%+-]+)@([A-Za-z0-9.-]+\.[A-Za-z]{2,})",
        lambda m: _mask_email_for_logs(m.group(0)),
        masked,
    )
    return masked


def _extract_detail_pairs(detail: str):
    raw = (detail or "").strip()
    if not raw:
        return []
    chunks = [c.strip() for c in raw.split(";") if c.strip()]
    out = []
    for ch in chunks:
        if "=" not in ch:
            out.append(("detalle", _mask_sensitive_text(ch)))
            continue
        k, v = ch.split("=", 1)
        out.append((k.strip().lower(), _mask_sensitive_text(v.strip())))
    return out


def _extract_request_id(detail: str):
    raw = (detail or "").strip()
    if not raw:
        return ""
    m = re.search(r"(?:^|[;\s,])rid=([A-Za-z0-9_-]{6,64})", raw)
    return m.group(1) if m else ""


def _extract_meta_from_detail(detail: str):
    pairs = dict(_extract_detail_pairs(detail))
    return {
        "rid": pairs.get("rid", ""),
        "method": pairs.get("method", ""),
        "path": pairs.get("path", ""),
        "endpoint": pairs.get("endpoint", ""),
    }


def _admin_log_module(action: str):
    key = (action or "").lower()
    if key.startswith("security_"):
        return "Seguridad", "admin_security_page"
    if key.startswith("user_"):
        return "Usuarios", "admin_users_page"
    if key.startswith("admin_"):
        return "Administradores", "admin_admins_page"
    return "Logs", "admin_logs_page"


def _admin_log_severity(action: str, detail: str = ""):
    key = (action or "").lower()
    d = (detail or "").lower()
    critical_words = ("delete", "revoke", "force_logout", "block")
    warn_words = ("suspend", "unlock", "clear", "status_change")
    if any(w in key for w in critical_words):
        return "critical", "Critico"
    if any(w in key for w in warn_words):
        return "warn", "Advertencia"
    if "error" in d or "fail" in d:
        return "warn", "Advertencia"
    return "info", "Info"


def _admin_filter_logs_rows(
    rows,
    q="",
    action="",
    actor="",
    ip="",
    target_user="",
    event_id="",
    request_id="",
    severity="",
    date_from=None,
    date_to=None,
):
    q = (q or "").strip().lower()
    action = (action or "").strip().lower()
    actor = (actor or "").strip().lower()
    ip = (ip or "").strip().lower()
    target_user = (target_user or "").strip().lower()
    event_id = (event_id or "").strip()
    request_id = (request_id or "").strip().lower()
    severity = (severity or "").strip().lower()

    filtered = []
    for log_row, actor_user in rows:
        actor_email = ((actor_user.email if actor_user else "") or "").lower()
        actor_name = ((actor_user.name if actor_user else "") or "").lower()
        action_val = (log_row.action or "").lower()
        detail_val = (log_row.detail or "").lower()
        ip_val = (log_row.ip or "").lower()
        rid_val = _extract_request_id(log_row.detail).lower()
        sev_key, _sev_label = _admin_log_severity(log_row.action, log_row.detail)
        created = to_naive_utc(log_row.created_at)

        if q:
            if not (
                q in actor_email
                or q in actor_name
                or q in action_val
                or q in detail_val
                or q in ip_val
                or q in str(log_row.id)
                or q in str(log_row.target_user_id or "")
                or q in rid_val
            ):
                continue
        if action and action not in action_val:
            continue
        if actor and actor not in actor_email and actor not in actor_name:
            continue
        if ip and ip not in ip_val:
            continue
        if target_user and target_user not in str(log_row.target_user_id or ""):
            continue
        if event_id and event_id != str(log_row.id):
            continue
        if request_id and request_id != rid_val:
            continue
        if severity and severity != sev_key:
            continue
        if date_from and (not created or created < date_from):
            continue
        if date_to and (not created or created >= date_to):
            continue
        filtered.append((log_row, actor_user))

    return filtered


def _admin_enrich_logs_rows(rows, now_utc=None):
    now_utc = now_utc or utcnow_naive()
    target_ids = {int(r[0].target_user_id) for r in rows if r[0].target_user_id}
    target_map = {}
    if target_ids:
        target_users = User.query.filter(User.id.in_(list(target_ids))).all()
        target_map = {u.id: u for u in target_users}

    enriched = []
    for log_row, actor_user in rows:
        title, icon, tone = _admin_action_meta(log_row.action)
        severity_key, severity_label = _admin_log_severity(log_row.action, log_row.detail or "")
        meta = _extract_meta_from_detail(log_row.detail or "")
        rid_val = _extract_request_id(log_row.detail or "")
        module_name, module_endpoint = _admin_log_module(log_row.action)
        module_url = url_for(module_endpoint)
        target_user = target_map.get(int(log_row.target_user_id or 0))
        detail_masked = _mask_sensitive_text(log_row.detail or "-")
        detail_pairs = _extract_detail_pairs(log_row.detail or "")

        summary = detail_masked
        if len(summary) > 150:
            summary = summary[:147] + "..."

        enriched.append({
            "log_row": log_row,
            "actor_user": actor_user,
            "target_user": target_user,
            "title": title,
            "icon": icon,
            "tone": tone,
            "severity_key": severity_key,
            "severity_label": severity_label,
            "when": _time_ago_es(log_row.created_at, now_utc),
            "summary": summary,
            "detail_masked": detail_masked,
            "detail_pairs": detail_pairs,
            "request_id": rid_val,
            "method": meta.get("method", ""),
            "path": meta.get("path", ""),
            "endpoint": meta.get("endpoint", ""),
            "module_name": module_name,
            "module_url": module_url,
        })
    return enriched


def _admin_logs_filters_from_request():
    filters = {
        "q": (request.args.get("q") or "").strip().lower(),
        "action": (request.args.get("action") or "").strip().lower(),
        "actor": (request.args.get("actor") or "").strip().lower(),
        "ip": (request.args.get("ip") or "").strip().lower(),
        "target_user": (request.args.get("target_user") or "").strip().lower(),
        "event_id": (request.args.get("event_id") or "").strip(),
        "request_id": (request.args.get("request_id") or "").strip().lower(),
        "severity": (request.args.get("severity") or "").strip().lower(),
        "date_from": _parse_date_ymd(request.args.get("date_from")),
        "date_to": _parse_date_ymd(request.args.get("date_to")),
    }
    if filters["date_to"]:
        filters["date_to"] = filters["date_to"] + timedelta(days=1)
    return filters


def _admin_logs_for_export(limit=5000):
    rows = _admin_recent_logs(limit=limit)
    filters = _admin_logs_filters_from_request()
    return _admin_filter_logs_rows(rows, **filters)


def _time_ago_es(dt, now_dt=None):
    dt_n = to_naive_utc(dt)
    now_n = to_naive_utc(now_dt or utcnow_naive())
    if not dt_n:
        return "-"
    if now_n < dt_n:
        return "Ahora"
    sec = int((now_n - dt_n).total_seconds())
    if sec < 60:
        return "Hace segundos"
    mins = sec // 60
    if mins < 60:
        return f"Hace {mins} min"
    hours = mins // 60
    if hours < 24:
        return f"Hace {hours} h"
    days = hours // 24
    return f"Hace {days} d"


def _admin_action_meta(action):
    action_key = (action or "").strip().lower()
    mapping = {
        "admin_grant": ("Permisos admin actualizados", "fa-user-shield", "ok"),
        "admin_revoke": ("Acceso admin revocado", "fa-user-slash", "warn"),
        "user_status_change": ("Estado de cuenta cambiado", "fa-user-gear", "warn"),
        "user_suspend": ("Cuenta suspendida", "fa-clock", "warn"),
        "user_unsuspend": ("Suspension retirada", "fa-unlock", "ok"),
        "user_delete": ("Cuenta eliminada", "fa-trash", "off"),
        "user_bulk_status_change": ("Cambio masivo de cuentas", "fa-users-gear", "warn"),
        "user_bulk_export": ("Exportacion masiva", "fa-file-export", "ok"),
    }
    return mapping.get(action_key, ("Evento administrativo", "fa-clipboard-list", "ok"))


def _admin_activity_feed(limit=12):
    rows = _admin_recent_logs(limit=limit)
    now_utc = utcnow_naive()
    items = []
    for log_row, actor_user in rows:
        title, icon, tone = _admin_action_meta(log_row.action)
        actor_name = "Sistema"
        if actor_user:
            actor_name = (actor_user.name or actor_user.email or "Sistema").strip()
        target = f"Usuario #{log_row.target_user_id}" if log_row.target_user_id else "Sin objetivo"
        detail = (log_row.detail or "").strip()
        if len(detail) > 160:
            detail = detail[:157] + "..."
        items.append({
            "title": title,
            "icon": icon,
            "tone": tone,
            "actor": actor_name,
            "target": target,
            "detail": detail or "-",
            "when": _time_ago_es(log_row.created_at, now_utc),
            "at": _format_dt_human(log_row.created_at),
        })
    return items


def _admin_alerts_payload():
    now_utc = utcnow_naive()
    items = []
    critical = 0
    warning = 0

    blocked_login = LoginAttempt.query.filter(
        LoginAttempt.blocked_until.isnot(None),
        LoginAttempt.blocked_until > now_utc
    ).count()
    if blocked_login > 0:
        items.append({
            "tone": "off",
            "icon": "fa-user-lock",
            "title": "Cuentas bloqueadas por intentos",
            "detail": f"Bloqueos activos en login: {blocked_login}",
        })
        critical += 1

    blocked_rate = RateLimit.query.filter(
        RateLimit.blocked_until.isnot(None),
        RateLimit.blocked_until > now_utc
    ).count()
    if blocked_rate > 0:
        items.append({
            "tone": "warn",
            "icon": "fa-gauge-high",
            "title": "Rate limits activos",
            "detail": f"Claves temporalmente bloqueadas: {blocked_rate}",
        })
        warning += 1

    blocked_reset_ip = ResetIPRequest.query.filter(
        ResetIPRequest.blocked_until.isnot(None),
        ResetIPRequest.blocked_until > now_utc
    ).count()
    if blocked_reset_ip > 0:
        items.append({
            "tone": "warn",
            "icon": "fa-shield-halved",
            "title": "Bloqueos en recuperacion por IP",
            "detail": f"IPs bloqueadas en reset: {blocked_reset_ip}",
        })
        warning += 1

    suspended_users = User.query.filter(
        User.suspended_until.isnot(None),
        User.suspended_until > now_utc
    ).count()
    if suspended_users > 0:
        items.append({
            "tone": "warn",
            "icon": "fa-clock",
            "title": "Cuentas suspendidas",
            "detail": f"Cuentas con suspension activa: {suspended_users}",
        })
        warning += 1

    recent_critical_events = AdminAuditLog.query.filter(
        AdminAuditLog.created_at >= (now_utc - timedelta(hours=12)),
        AdminAuditLog.action.in_(["user_delete", "admin_revoke"])
    ).count()
    if recent_critical_events > 0:
        items.append({
            "tone": "off",
            "icon": "fa-triangle-exclamation",
            "title": "Eventos administrativos criticos",
            "detail": f"Eventos en ultimas 12h: {recent_critical_events}",
        })
        critical += 1

    gemini_count = len([k for k in LISTA_DE_CLAVES if k])
    if gemini_count == 0:
        items.append({
            "tone": "off",
            "icon": "fa-robot",
            "title": "Gemini sin claves",
            "detail": "No hay claves disponibles para responder.",
        })
        critical += 1

    if not items:
        items.append({
            "tone": "ok",
            "icon": "fa-circle-check",
            "title": "Sin alertas criticas",
            "detail": "El sistema se mantiene estable en este momento.",
        })

    return {
        "critical_count": critical,
        "warning_count": warning,
        "items": items,
        "generated_at": _format_dt_human(now_utc),
    }


def _admin_security_data(limit=100):
    login_attempts = LoginAttempt.query.order_by(LoginAttempt.id.desc()).limit(limit).all()
    reset_ip_rows = ResetIPRequest.query.order_by(ResetIPRequest.id.desc()).limit(limit).all()
    rate_limit_rows = RateLimit.query.order_by(RateLimit.id.desc()).limit(limit).all()
    return login_attempts, reset_ip_rows, rate_limit_rows


def _security_can_manage_actions(user: User) -> bool:
    role = _effective_admin_role(user)
    perms = _effective_admin_permissions(user)
    return bool(
        role == "super_admin"
        or "manage_users" in perms
        or "manage_settings" in perms
    )


def _security_block_state(blocked_until, now_utc=None):
    now_utc = now_utc or utcnow_naive()
    until = to_naive_utc(blocked_until)
    is_blocked = bool(until and until > now_utc)
    return is_blocked, until


def _admin_security_redirect():
    base = url_for("admin_security_page")
    candidate = (request.form.get("return_to") or request.referrer or "").strip()
    if not candidate:
        return redirect(base)
    if candidate.startswith(base):
        return redirect(candidate)
    try:
        parsed = urlparse(candidate)
    except Exception:
        parsed = None
    if parsed and parsed.path == base:
        safe = parsed.path + (f"?{parsed.query}" if parsed.query else "")
        return redirect(safe)
    return redirect(base)


def _security_duration_delta(value_raw, unit_raw):
    value = _safe_int(value_raw, 0)
    unit = (unit_raw or "").strip().lower()
    if value < 1:
        return None
    if unit == "minutes":
        return timedelta(minutes=value)
    if unit == "hours":
        return timedelta(hours=value)
    if unit == "days":
        return timedelta(days=value)
    return None


def _safe_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_date_ymd(value: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


def _slice_with_pagination(items, page, per_page):
    total = len(items)
    total_pages = max(1, math.ceil(total / per_page)) if per_page > 0 else 1
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    return items[start:end], total, total_pages, page


def _build_pagination_links(endpoint, args_dict, page_key, per_key, page, per_page, total_pages):
    pages = []
    if total_pages <= 1:
        return {"pages": pages, "prev_url": None, "next_url": None}

    window = 2
    start = max(1, page - window)
    end = min(total_pages, page + window)
    for num in range(start, end + 1):
        params = dict(args_dict)
        params[page_key] = num
        params[per_key] = per_page
        pages.append({
            "num": num,
            "current": (num == page),
            "url": f"{url_for(endpoint)}?{urlencode(params)}",
        })

    prev_url = None
    next_url = None
    if page > 1:
        p = dict(args_dict)
        p[page_key] = page - 1
        p[per_key] = per_page
        prev_url = f"{url_for(endpoint)}?{urlencode(p)}"
    if page < total_pages:
        p = dict(args_dict)
        p[page_key] = page + 1
        p[per_key] = per_page
        next_url = f"{url_for(endpoint)}?{urlencode(p)}"

    return {"pages": pages, "prev_url": prev_url, "next_url": next_url}


@app.route('/admin')
@login_required
@admin_required(permission="view_dashboard")
def admin_panel():
    role = _effective_admin_role(current_user)
    perms = _effective_admin_permissions(current_user)

    stats = _admin_stats()
    charts = _admin_dashboard_charts(days=7)
    system_health = _admin_system_health()
    activity_feed = _admin_activity_feed(limit=12)
    alerts_payload = _admin_alerts_payload()
    top_chat_users = (
        db.session.query(
            User.id,
            User.name,
            User.email,
            func.count(Conversation.id).label("chat_count"),
        )
        .outerjoin(Conversation, Conversation.user_id == User.id)
        .group_by(User.id, User.name, User.email)
        .order_by(func.count(Conversation.id).desc(), User.id.asc())
        .limit(8)
        .all()
    )
    top_chat_labels = [r.name for r in top_chat_users]
    top_chat_values = [int(r.chat_count or 0) for r in top_chat_users]

    return render_template(
        'admin_panel.html',
        admin_role=role,
        admin_permissions=sorted(perms, key=lambda p: _permission_label_es(p)),
        total_users=stats["total_users"],
        total_chats=stats["total_chats"],
        total_messages=stats["total_messages"],
        total_admins=stats["total_admins"],
        top_chat_labels=top_chat_labels,
        top_chat_values=top_chat_values,
        message_day_labels=charts["message_day_labels"],
        message_day_values=charts["message_day_values"],
        account_status_labels=charts["account_status_labels"],
        account_status_values=charts["account_status_values"],
        failed_ip_labels=charts["failed_ip_labels"],
        failed_ip_values=charts["failed_ip_values"],
        system_health=system_health,
        activity_feed=activity_feed,
        alerts_payload=alerts_payload,
        can_view_logs=("view_logs" in perms),
    )


@app.route('/admin/alerts_feed')
@login_required
@admin_required(permission="view_dashboard")
def admin_alerts_feed():
    payload = _admin_alerts_payload()
    return jsonify({"success": True, **payload})


@app.route('/admin/admins')
@login_required
@admin_required(super_only=True)
def admin_admins_page():
    role = _effective_admin_role(current_user)
    perms = _effective_admin_permissions(current_user)
    stats = _admin_stats()
    args = request.args.to_dict(flat=True)
    per_options = [5, 10, 20, 50, 100]

    admins_data_all = _admin_admins_data()
    admin_last_activity_rows = (
        db.session.query(
            AdminAuditLog.actor_user_id.label("uid"),
            func.max(AdminAuditLog.created_at).label("last_at"),
        )
        .filter(AdminAuditLog.actor_user_id.isnot(None))
        .group_by(AdminAuditLog.actor_user_id)
        .all()
    )
    admin_last_activity_map = {int(r.uid): r.last_at for r in admin_last_activity_rows if r.uid}
    for row in admins_data_all:
        uid = int(row["user_row"].id)
        last_at = admin_last_activity_map.get(uid)
        row["last_activity_at"] = last_at
        row["last_activity_ago"] = _time_ago_es(last_at) if last_at else "Sin actividad"
        row["role_label"] = "Super Admin" if row["role_row"].role == "super_admin" else "Admin"
        row["is_protected"] = bool(row["role_row"].role == "super_admin")

    admins_super_total = sum(1 for row in admins_data_all if row["role_row"].role == "super_admin")
    admins_standard_total = sum(1 for row in admins_data_all if row["role_row"].role == "admin")
    admins_active_total = sum(1 for row in admins_data_all if bool(row["role_row"].is_active))
    admins_inactive_total = max(0, len(admins_data_all) - admins_active_total)
    admins_with_activity_total = sum(1 for row in admins_data_all if row["last_activity_at"])
    admins_recent_activity = _admin_activity_feed(limit=4)
    admins_q = (request.args.get("admins_q") or "").strip().lower()
    admins_role = (request.args.get("admins_role") or "").strip().lower()
    admins_state = (request.args.get("admins_state") or "").strip().lower()

    if admins_q:
        admins_data_all = [
            row for row in admins_data_all
            if admins_q in (row["user_row"].name or "").lower()
            or admins_q in (row["user_row"].email or "").lower()
            or admins_q == str(row["user_row"].id)
        ]
    if admins_role in {"admin", "super_admin"}:
        admins_data_all = [row for row in admins_data_all if row["role_row"].role == admins_role]
    if admins_state in {"activo", "inactivo"}:
        is_active = admins_state == "activo"
        admins_data_all = [row for row in admins_data_all if bool(row["role_row"].is_active) == is_active]

    per_admins = _safe_int(request.args.get("per_admins"), 10)
    if per_admins not in per_options:
        per_admins = 10
    page_admins = _safe_int(request.args.get("page_admins"), 1)
    admins_data, admins_total, admins_total_pages, page_admins = _slice_with_pagination(
        admins_data_all, page_admins, per_admins
    )
    admins_pagination = _build_pagination_links(
        "admin_admins_page", args, "page_admins", "per_admins", page_admins, per_admins, admins_total_pages
    )

    logs_all = _admin_recent_logs(limit=3000)
    logs_q = (request.args.get("logs_q") or "").strip().lower()
    logs_actor = (request.args.get("logs_actor") or "").strip().lower()
    logs_action = (request.args.get("logs_action") or "").strip().lower()
    logs_date_from = _parse_date_ymd(request.args.get("logs_date_from"))
    logs_date_to = _parse_date_ymd(request.args.get("logs_date_to"))
    if logs_date_to:
        logs_date_to = logs_date_to + timedelta(days=1)
    if logs_q:
        logs_all = [
            row for row in logs_all
            if logs_q in ((row[1].email if row[1] else "") or "").lower()
            or logs_q in (row[0].action or "").lower()
            or logs_q in (row[0].detail or "").lower()
            or logs_q == str(row[0].id)
            or logs_q == str(row[0].target_user_id or "")
        ]
    if logs_actor:
        logs_all = [
            row for row in logs_all
            if logs_actor in ((row[1].email if row[1] else "") or "").lower()
            or logs_actor in ((row[1].name if row[1] else "") or "").lower()
        ]
    if logs_action:
        logs_all = [row for row in logs_all if logs_action in (row[0].action or "").lower()]
    if logs_date_from:
        logs_all = [row for row in logs_all if row[0].created_at and row[0].created_at >= logs_date_from]
    if logs_date_to:
        logs_all = [row for row in logs_all if row[0].created_at and row[0].created_at < logs_date_to]

    per_logs = _safe_int(request.args.get("per_logs"), 20)
    if per_logs not in per_options:
        per_logs = 20
    page_logs = _safe_int(request.args.get("page_logs"), 1)
    recent_admin_logs, logs_total, logs_total_pages, page_logs = _slice_with_pagination(
        logs_all, page_logs, per_logs
    )
    logs_pagination = _build_pagination_links(
        "admin_admins_page", args, "page_logs", "per_logs", page_logs, per_logs, logs_total_pages
    )
    logs_timeline = []
    for log_row, actor_user in recent_admin_logs:
        title, icon, tone = _admin_action_meta(log_row.action)
        logs_timeline.append({
            "log_row": log_row,
            "actor_user": actor_user,
            "title": title,
            "icon": icon,
            "tone": tone,
            "when": _time_ago_es(log_row.created_at),
        })
    logs_action_options = sorted({
        (row[0].action or "").strip()
        for row in logs_all
        if (row[0].action or "").strip()
    })

    return render_template(
        'admin_admins.html',
        admin_role=role,
        admin_permissions=sorted(perms, key=lambda p: _permission_label_es(p)),
        total_users=stats["total_users"],
        total_chats=stats["total_chats"],
        total_messages=stats["total_messages"],
        total_admins=stats["total_admins"],
        admins_data=admins_data,
        recent_admin_logs=recent_admin_logs,
        logs_timeline=logs_timeline,
        admins_q=admins_q,
        admins_role=admins_role,
        admins_state=admins_state,
        logs_q=logs_q,
        logs_actor=logs_actor,
        logs_action=logs_action,
        logs_date_from=(request.args.get("logs_date_from") or ""),
        logs_date_to=(request.args.get("logs_date_to") or ""),
        logs_action_options=logs_action_options,
        per_options=per_options,
        per_admins=per_admins,
        per_logs=per_logs,
        admins_total=admins_total,
        logs_total=logs_total,
        admins_super_total=admins_super_total,
        admins_standard_total=admins_standard_total,
        admins_active_total=admins_active_total,
        admins_inactive_total=admins_inactive_total,
        admins_with_activity_total=admins_with_activity_total,
        admins_recent_activity=admins_recent_activity,
        admins_pagination=admins_pagination,
        logs_pagination=logs_pagination,
        all_permissions=sorted(ALL_ADMIN_PERMISSIONS, key=lambda p: _permission_label_es(p)),
        default_permissions=sorted(DEFAULT_ADMIN_PERMISSIONS, key=lambda p: _permission_label_es(p)),
        permission_labels_es=PERMISSION_LABELS_ES,
        permission_groups_es=PERMISSION_GROUPS_ES,
        super_admin_emails=SUPER_ADMIN_EMAILS,
    )


@app.route('/admin/usuarios')
@login_required
@admin_required(permission="view_users")
def admin_users_page():
    role = _effective_admin_role(current_user)
    perms = _effective_admin_permissions(current_user)
    stats = _admin_stats()
    args = request.args.to_dict(flat=True)
    per_options = [5, 10, 20, 50, 100]
    now_utc = utcnow_naive()

    users_all = _admin_users_data()
    q = (request.args.get("q") or "").strip().lower()
    estado = (request.args.get("estado") or "").strip().lower()
    date_from = _parse_date_ymd(request.args.get("date_from"))
    date_to = _parse_date_ymd(request.args.get("date_to"))
    if date_to:
        date_to = date_to + timedelta(days=1)

    if q:
        users_all = [
            u for u in users_all
            if q in (u.name or "").lower() or q in (u.email or "").lower() or q == str(u.id)
        ]
    if estado in {"activa", "desactivada", "suspendida"}:
        users_all = [u for u in users_all if _user_status_data(u, now_utc)[0] == estado]
    if date_from:
        users_all = [u for u in users_all if u.created_at and u.created_at >= date_from]
    if date_to:
        users_all = [u for u in users_all if u.created_at and u.created_at < date_to]

    per_page = _safe_int(request.args.get("per_page"), 10)
    if per_page not in per_options:
        per_page = 10
    page = _safe_int(request.args.get("page"), 1)
    all_users, users_total, users_total_pages, page = _slice_with_pagination(users_all, page, per_page)
    users_pagination = _build_pagination_links(
        "admin_users_page", args, "page", "per_page", page, per_page, users_total_pages
    )

    return render_template(
        'admin_users.html',
        admin_role=role,
        super_admin_emails=SUPER_ADMIN_EMAILS,
        admin_permissions=sorted(perms, key=lambda p: _permission_label_es(p)),
        can_manage_users=("manage_users" in perms),
        can_export_reports=("export_reports" in perms),
        is_super_admin=(role == "super_admin"),
        total_users=stats["total_users"],
        total_chats=stats["total_chats"],
        total_messages=stats["total_messages"],
        total_admins=stats["total_admins"],
        all_users=all_users,
        q=q,
        estado=estado,
        now_utc=now_utc,
        date_from=(request.args.get("date_from") or ""),
        date_to=(request.args.get("date_to") or ""),
        per_page=per_page,
        per_options=per_options,
        users_total=users_total,
        users_pagination=users_pagination,
    )


@app.route('/admin/logs')
@login_required
@admin_required(permission="view_logs")
def admin_logs_page():
    role = _effective_admin_role(current_user)
    perms = _effective_admin_permissions(current_user)
    stats = _admin_stats()
    now_utc = utcnow_naive()
    args = request.args.to_dict(flat=True)
    per_options = [5, 10, 20, 50, 100]
    logs_all = _admin_recent_logs(limit=5000)
    q = (request.args.get("q") or "").strip().lower()
    action = (request.args.get("action") or "").strip().lower()
    actor = (request.args.get("actor") or "").strip().lower()
    ip = (request.args.get("ip") or "").strip().lower()
    target_user = (request.args.get("target_user") or "").strip().lower()
    event_id = (request.args.get("event_id") or "").strip()
    request_id = (request.args.get("request_id") or "").strip().lower()
    severity = (request.args.get("severity") or "").strip().lower()
    date_from = _parse_date_ymd(request.args.get("date_from"))
    date_to = _parse_date_ymd(request.args.get("date_to"))
    if date_to:
        date_to = date_to + timedelta(days=1)

    logs_action_options = sorted({
        (row[0].action or "").strip()
        for row in logs_all
        if (row[0].action or "").strip()
    })
    logs_actor_options = sorted({
        ((row[1].email if row[1] else "") or "").strip()
        for row in logs_all
        if ((row[1].email if row[1] else "") or "").strip()
    })[:120]

    filtered_logs = _admin_filter_logs_rows(
        logs_all,
        q=q,
        action=action,
        actor=actor,
        ip=ip,
        target_user=target_user,
        event_id=event_id,
        request_id=request_id,
        severity=severity,
        date_from=date_from,
        date_to=date_to,
    )

    enriched_all = _admin_enrich_logs_rows(filtered_logs, now_utc)
    logs_total = len(filtered_logs)

    events_today = 0
    critical_today = 0
    actors_unique = set()
    ip_counter = Counter()
    for item in enriched_all:
        created = to_naive_utc(item["log_row"].created_at)
        if created and created.date() == now_utc.date():
            events_today += 1
            if item["severity_key"] == "critical":
                critical_today += 1
        actor_email = ((item["actor_user"].email if item["actor_user"] else "") or "").strip().lower()
        if actor_email:
            actors_unique.add(actor_email)
        ip_val = (item["log_row"].ip or "").strip()
        if ip_val:
            ip_counter[ip_val] += 1
    suspicious_ips = len([ip_key for ip_key, cnt in ip_counter.items() if cnt >= 3])
    timeline_items = enriched_all[:12]

    per_page = _safe_int(request.args.get("per_page"), 20)
    if per_page not in per_options:
        per_page = 20
    page = _safe_int(request.args.get("page"), 1)
    paginated_rows, logs_total, logs_total_pages, page = _slice_with_pagination(filtered_logs, page, per_page)
    logs_pagination = _build_pagination_links(
        "admin_logs_page", args, "page", "per_page", page, per_page, logs_total_pages
    )
    table_rows = _admin_enrich_logs_rows(paginated_rows, now_utc)

    export_args = {}
    for key in ["q", "action", "actor", "ip", "target_user", "event_id", "request_id", "severity", "date_from", "date_to"]:
        val = request.args.get(key)
        if val:
            export_args[key] = val
    export_query = urlencode(export_args)

    return render_template(
        'admin_logs.html',
        admin_role=role,
        admin_permissions=sorted(perms, key=lambda p: _permission_label_es(p)),
        total_users=stats["total_users"],
        total_chats=stats["total_chats"],
        total_messages=stats["total_messages"],
        total_admins=stats["total_admins"],
        timeline_items=timeline_items,
        table_rows=table_rows,
        q=q,
        action=action,
        actor=actor,
        ip=ip,
        target_user=target_user,
        event_id=event_id,
        request_id=request_id,
        severity=severity,
        logs_action_options=logs_action_options,
        logs_actor_options=logs_actor_options,
        date_from=(request.args.get("date_from") or ""),
        date_to=(request.args.get("date_to") or ""),
        per_page=per_page,
        per_options=per_options,
        logs_total=logs_total,
        events_today=events_today,
        critical_today=critical_today,
        actors_unique=len(actors_unique),
        suspicious_ips=suspicious_ips,
        export_query=export_query,
        can_cleanup_logs=(role == "super_admin"),
        retention_days=ADMIN_LOG_RETENTION_DAYS,
        logs_pagination=logs_pagination,
    )


@app.route('/admin/logs/cleanup', methods=['POST'])
@login_required
@admin_required(super_only=True)
def admin_logs_cleanup():
    deleted = _cleanup_old_admin_logs(force=True)
    if deleted > 0:
        _add_admin_audit(
            "logs_cleanup",
            detail=f"deleted={deleted}; retention_days={ADMIN_LOG_RETENTION_DAYS}",
        )
        flash(f"Limpieza aplicada: {deleted} logs eliminados por retencion.", "success")
    else:
        flash("No habia logs antiguos para limpiar.", "warning")
    return redirect(url_for('admin_logs_page'))


@app.route('/admin/seguridad')
@login_required
@admin_required(permission="view_security")
def admin_security_page():
    role = _effective_admin_role(current_user)
    perms = _effective_admin_permissions(current_user)
    stats = _admin_stats()
    now_utc = utcnow_naive()
    can_security_actions = _security_can_manage_actions(current_user)
    args = request.args.to_dict(flat=True)
    per_options = [5, 10, 20, 50, 100]
    login_attempts_all, reset_ip_all, rate_limit_all = _admin_security_data(limit=5000)
    security_blocks_all = (
        SecurityBlock.query.filter_by(is_active=True)
        .filter(SecurityBlock.blocked_until > now_utc)
        .order_by(SecurityBlock.blocked_until.desc(), SecurityBlock.id.desc())
        .limit(5000)
        .all()
    )

    login_blocked_active = 0
    reset_blocked_active = 0
    rate_blocked_active = 0
    login_risky_rows = 0
    recent_lock_events = 0

    for row in login_attempts_all:
        is_blocked, until = _security_block_state(row.blocked_until, now_utc)
        attempts = int(row.attempts or 0)
        row.status_key = "off" if is_blocked else ("warn" if attempts > 0 else "ok")
        row.status_label = "Bloqueado" if is_blocked else ("En observacion" if attempts > 0 else "Normal")
        row.blocked_until_human = _format_dt_human(until) if until else "-"
        row.can_unlock = bool(is_blocked or attempts > 0)
        if is_blocked:
            login_blocked_active += 1
        if attempts >= max(3, LOGIN_MAX_ATTEMPTS - 2):
            login_risky_rows += 1
        first_at = to_naive_utc(row.first_attempt_at)
        if first_at and (now_utc - first_at) <= timedelta(hours=24) and is_blocked:
            recent_lock_events += 1

    for row in reset_ip_all:
        is_blocked, until = _security_block_state(row.blocked_until, now_utc)
        attempts = int(row.attempts or 0)
        row.status_key = "off" if is_blocked else ("warn" if attempts > 0 else "ok")
        row.status_label = "Bloqueado" if is_blocked else ("En observacion" if attempts > 0 else "Normal")
        row.blocked_until_human = _format_dt_human(until) if until else "-"
        row.can_unlock = bool(is_blocked or attempts > 0)
        if is_blocked:
            reset_blocked_active += 1

    for row in rate_limit_all:
        is_blocked, until = _security_block_state(row.blocked_until, now_utc)
        count = int(row.count or 0)
        row.status_key = "off" if is_blocked else ("warn" if count > 0 else "ok")
        row.status_label = "Bloqueado" if is_blocked else ("Activo" if count > 0 else "Limpio")
        row.blocked_until_human = _format_dt_human(until) if until else "-"
        row.can_unlock = bool(is_blocked or count > 0)
        if is_blocked:
            rate_blocked_active += 1

    manual_block_active = len(security_blocks_all)
    total_active_blocks = login_blocked_active + reset_blocked_active + rate_blocked_active + manual_block_active
    if total_active_blocks >= 10:
        security_risk_label = "Riesgo alto"
        security_risk_tone = "off"
    elif total_active_blocks >= 3:
        security_risk_label = "Riesgo medio"
        security_risk_tone = "warn"
    else:
        security_risk_label = "Riesgo controlado"
        security_risk_tone = "ok"

    login_q = (request.args.get("login_q") or "").strip().lower()
    if login_q:
        login_attempts_all = [
            r for r in login_attempts_all
            if login_q in (r.ip or "").lower() or login_q in (r.email or "").lower() or login_q == str(r.id)
        ]
    reset_q = (request.args.get("reset_q") or "").strip().lower()
    if reset_q:
        reset_ip_all = [
            r for r in reset_ip_all
            if reset_q in (r.ip or "").lower() or reset_q == str(r.id)
        ]
    rate_q = (request.args.get("rate_q") or "").strip().lower()
    if rate_q:
        rate_limit_all = [
            r for r in rate_limit_all
            if rate_q in (r.key or "").lower() or rate_q == str(r.id)
        ]

    block_q = (request.args.get("block_q") or "").strip().lower()
    if block_q:
        security_blocks_all = [
            b for b in security_blocks_all
            if block_q in (b.target or "").lower()
            or block_q in (b.reason or "").lower()
            or block_q in (b.block_type or "").lower()
            or block_q == str(b.id)
        ]
    for b in security_blocks_all:
        b.type_label = "Correo" if b.block_type == "email" else "IP"
        b.until_human = _format_dt_human(b.blocked_until)
        b.status_key = "off"

    per_login = _safe_int(request.args.get("per_login"), 10)
    if per_login not in per_options:
        per_login = 10
    page_login = _safe_int(request.args.get("page_login"), 1)
    login_attempts, login_total, login_total_pages, page_login = _slice_with_pagination(
        login_attempts_all, page_login, per_login
    )
    login_pagination = _build_pagination_links(
        "admin_security_page", args, "page_login", "per_login", page_login, per_login, login_total_pages
    )

    per_reset = _safe_int(request.args.get("per_reset"), 10)
    if per_reset not in per_options:
        per_reset = 10
    page_reset = _safe_int(request.args.get("page_reset"), 1)
    reset_ip_rows, reset_total, reset_total_pages, page_reset = _slice_with_pagination(
        reset_ip_all, page_reset, per_reset
    )
    reset_pagination = _build_pagination_links(
        "admin_security_page", args, "page_reset", "per_reset", page_reset, per_reset, reset_total_pages
    )

    per_rate = _safe_int(request.args.get("per_rate"), 10)
    if per_rate not in per_options:
        per_rate = 10
    page_rate = _safe_int(request.args.get("page_rate"), 1)
    rate_limit_rows, rate_total, rate_total_pages, page_rate = _slice_with_pagination(
        rate_limit_all, page_rate, per_rate
    )
    rate_pagination = _build_pagination_links(
        "admin_security_page", args, "page_rate", "per_rate", page_rate, per_rate, rate_total_pages
    )

    per_block = _safe_int(request.args.get("per_block"), 10)
    if per_block not in per_options:
        per_block = 10
    page_block = _safe_int(request.args.get("page_block"), 1)
    security_blocks, block_total, block_total_pages, page_block = _slice_with_pagination(
        security_blocks_all, page_block, per_block
    )
    block_pagination = _build_pagination_links(
        "admin_security_page", args, "page_block", "per_block", page_block, per_block, block_total_pages
    )

    return render_template(
        'admin_security.html',
        admin_role=role,
        admin_permissions=sorted(perms, key=lambda p: _permission_label_es(p)),
        can_security_actions=can_security_actions,
        total_users=stats["total_users"],
        total_chats=stats["total_chats"],
        total_messages=stats["total_messages"],
        total_admins=stats["total_admins"],
        security_risk_label=security_risk_label,
        security_risk_tone=security_risk_tone,
        total_active_blocks=total_active_blocks,
        login_blocked_active=login_blocked_active,
        reset_blocked_active=reset_blocked_active,
        rate_blocked_active=rate_blocked_active,
        manual_block_active=manual_block_active,
        login_risky_rows=login_risky_rows,
        recent_lock_events=recent_lock_events,
        login_attempts=login_attempts,
        reset_ip_rows=reset_ip_rows,
        rate_limit_rows=rate_limit_rows,
        security_blocks=security_blocks,
        login_q=login_q,
        reset_q=reset_q,
        rate_q=rate_q,
        block_q=block_q,
        per_options=per_options,
        per_login=per_login,
        per_reset=per_reset,
        per_rate=per_rate,
        per_block=per_block,
        login_total=login_total,
        reset_total=reset_total,
        rate_total=rate_total,
        block_total=block_total,
        login_pagination=login_pagination,
        reset_pagination=reset_pagination,
        rate_pagination=rate_pagination,
        block_pagination=block_pagination,
    )


@app.route('/admin/seguridad/login/unlock/<int:row_id>', methods=['POST'])
@login_required
@admin_required(permission="view_security")
def admin_security_unlock_login(row_id):
    if not _security_can_manage_actions(current_user):
        flash("No tienes permisos para ejecutar acciones de seguridad.", "error")
        return _admin_security_redirect()

    row = db.session.get(LoginAttempt, row_id)
    if not row:
        flash("Registro de login no encontrado.", "error")
        return _admin_security_redirect()

    attempts_before = int(row.attempts or 0)
    blocked_before = _format_dt_human(row.blocked_until)
    row.attempts = 0
    row.first_attempt_at = None
    row.blocked_until = None
    db.session.commit()

    _add_admin_audit(
        "security_login_unlock",
        detail=(
            f"id={row.id}; email={row.email or '-'}; ip={row.ip or '-'}; "
            f"attempts_before={attempts_before}; blocked_until_before={blocked_before}"
        ),
    )
    flash("Login desbloqueado y contador reiniciado.", "success")
    return _admin_security_redirect()


@app.route('/admin/seguridad/reset_ip/unlock/<int:row_id>', methods=['POST'])
@login_required
@admin_required(permission="view_security")
def admin_security_unlock_reset_ip(row_id):
    if not _security_can_manage_actions(current_user):
        flash("No tienes permisos para ejecutar acciones de seguridad.", "error")
        return _admin_security_redirect()

    row = db.session.get(ResetIPRequest, row_id)
    if not row:
        flash("Registro de reset por IP no encontrado.", "error")
        return _admin_security_redirect()

    attempts_before = int(row.attempts or 0)
    blocked_before = _format_dt_human(row.blocked_until)
    row.attempts = 0
    row.first_attempt_at = None
    row.last_sent_at = None
    row.blocked_until = None
    db.session.commit()

    _add_admin_audit(
        "security_reset_ip_unlock",
        detail=(
            f"id={row.id}; ip={row.ip or '-'}; attempts_before={attempts_before}; "
            f"blocked_until_before={blocked_before}"
        ),
    )
    flash("Registro de reset por IP desbloqueado.", "success")
    return _admin_security_redirect()


@app.route('/admin/seguridad/rate_limit/clear/<int:row_id>', methods=['POST'])
@login_required
@admin_required(permission="view_security")
def admin_security_clear_rate_limit(row_id):
    if not _security_can_manage_actions(current_user):
        flash("No tienes permisos para ejecutar acciones de seguridad.", "error")
        return _admin_security_redirect()

    row = db.session.get(RateLimit, row_id)
    if not row:
        flash("Registro de rate limit no encontrado.", "error")
        return _admin_security_redirect()

    count_before = int(row.count or 0)
    blocked_before = _format_dt_human(row.blocked_until)
    row.count = 0
    row.window_start = utcnow_naive()
    row.blocked_until = None
    db.session.commit()

    _add_admin_audit(
        "security_rate_limit_clear",
        detail=(
            f"id={row.id}; key={row.key or '-'}; count_before={count_before}; "
            f"blocked_until_before={blocked_before}"
        ),
    )
    flash("Rate limit reiniciado correctamente.", "success")
    return _admin_security_redirect()


@app.route('/admin/seguridad/block/email', methods=['POST'])
@login_required
@admin_required(permission="view_security")
def admin_security_block_email():
    if not _security_can_manage_actions(current_user):
        flash("No tienes permisos para ejecutar acciones de seguridad.", "error")
        return _admin_security_redirect()

    email = _normalize_email(request.form.get("email"))
    reason = (request.form.get("reason") or "").strip()
    delta = _security_duration_delta(request.form.get("duration_value"), request.form.get("duration_unit"))

    if not email or not EMAIL_RE.match(email):
        flash("Debes ingresar un correo valido para bloquear.", "error")
        return _admin_security_redirect()
    if not delta:
        flash("Duracion invalida para bloqueo de correo.", "error")
        return _admin_security_redirect()

    until = utcnow_naive() + delta
    row = SecurityBlock.query.filter_by(block_type="email", target=email).first()
    if not row:
        row = SecurityBlock(
            block_type="email",
            target=email,
            created_by_user_id=current_user.id,
        )
        db.session.add(row)
    row.reason = reason or "Bloqueo manual de correo"
    row.blocked_until = until
    row.is_active = True
    row.created_by_user_id = current_user.id
    row.updated_at = utcnow_naive()
    db.session.commit()

    _add_admin_audit(
        "security_block_email_set",
        detail=f"email={email}; until={_format_dt_human(until)}; reason={row.reason}",
    )
    flash(f"Bloqueo por correo activo hasta {_format_dt_human(until)}.", "success")
    return _admin_security_redirect()


@app.route('/admin/seguridad/block/ip', methods=['POST'])
@login_required
@admin_required(permission="view_security")
def admin_security_block_ip():
    if not _security_can_manage_actions(current_user):
        flash("No tienes permisos para ejecutar acciones de seguridad.", "error")
        return _admin_security_redirect()

    ip = _normalize_ip(request.form.get("ip"))
    reason = (request.form.get("reason") or "").strip()
    delta = _security_duration_delta(request.form.get("duration_value"), request.form.get("duration_unit"))

    if not ip:
        flash("Debes ingresar una IP valida para bloquear.", "error")
        return _admin_security_redirect()
    if not delta:
        flash("Duracion invalida para bloqueo de IP.", "error")
        return _admin_security_redirect()

    until = utcnow_naive() + delta
    row = SecurityBlock.query.filter_by(block_type="ip", target=ip).first()
    if not row:
        row = SecurityBlock(
            block_type="ip",
            target=ip,
            created_by_user_id=current_user.id,
        )
        db.session.add(row)
    row.reason = reason or "Bloqueo manual de IP"
    row.blocked_until = until
    row.is_active = True
    row.created_by_user_id = current_user.id
    row.updated_at = utcnow_naive()
    db.session.commit()

    _add_admin_audit(
        "security_block_ip_set",
        detail=f"ip={ip}; until={_format_dt_human(until)}; reason={row.reason}",
    )
    flash(f"Bloqueo por IP activo hasta {_format_dt_human(until)}.", "success")
    return _admin_security_redirect()


@app.route('/admin/seguridad/block/remove/<int:block_id>', methods=['POST'])
@login_required
@admin_required(permission="view_security")
def admin_security_remove_block(block_id):
    if not _security_can_manage_actions(current_user):
        flash("No tienes permisos para ejecutar acciones de seguridad.", "error")
        return _admin_security_redirect()

    row = db.session.get(SecurityBlock, block_id)
    if not row:
        flash("Bloqueo no encontrado.", "error")
        return _admin_security_redirect()

    row.is_active = False
    row.updated_at = utcnow_naive()
    db.session.commit()

    _add_admin_audit(
        "security_block_remove",
        detail=f"id={row.id}; type={row.block_type}; target={row.target}",
    )
    flash("Bloqueo manual retirado.", "success")
    return _admin_security_redirect()


@app.route('/admin/seguridad/force_logout', methods=['POST'])
@login_required
@admin_required(permission="view_security")
def admin_security_force_logout():
    if not _security_can_manage_actions(current_user):
        flash("No tienes permisos para ejecutar acciones de seguridad.", "error")
        return _admin_security_redirect()

    email = _normalize_email(request.form.get("email"))
    if not email:
        flash("Debes ingresar un correo para cierre forzado.", "error")
        return _admin_security_redirect()

    user = User.query.filter(func.lower(User.email) == email).first()
    if not user:
        flash("No existe un usuario con ese correo.", "error")
        return _admin_security_redirect()

    _mark_force_logout(user.id)
    _add_admin_audit(
        "security_force_logout",
        target_user_id=user.id,
        detail=f"email={user.email}",
    )
    flash(f"Se marco cierre forzado de sesion para {user.email}.", "success")
    return _admin_security_redirect()


@app.route('/admin/reportes')
@login_required
@admin_required(permission="export_reports")
def admin_reports_page():
    role = _effective_admin_role(current_user)
    perms = _effective_admin_permissions(current_user)
    stats = _admin_stats()
    return render_template(
        'admin_reports.html',
        admin_role=role,
        admin_permissions=sorted(perms, key=lambda p: _permission_label_es(p)),
        total_users=stats["total_users"],
        total_chats=stats["total_chats"],
        total_messages=stats["total_messages"],
        total_admins=stats["total_admins"],
    )


def _build_xlsx_response(filename_base: str, sheet_name: str, headers: list, rows: list):
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Reporte")[:31]

    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    header_fill = PatternFill(start_color="13233F", end_color="13233F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center

    for idx, _ in enumerate(headers, start=1):
        max_len = len(str(headers[idx - 1]))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=idx).value
            val_len = len(str(val)) if val is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 52)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"{filename_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route('/admin/export/usuarios.xlsx')
@login_required
@admin_required(permission="export_reports")
def admin_export_users_xlsx():
    rows = _admin_users_data()
    headers = ["ID", "Nombre", "Email", "Registro", "Estado", "Suspendida hasta", "Chats", "Mensajes"]
    values = []
    now_utc = utcnow_naive()
    for u in rows:
        status_key, status_label, suspended_until = _user_status_data(u, now_utc)
        values.append([
            u.id,
            u.name or "",
            u.email or "",
            u.created_at.strftime('%Y-%m-%d %H:%M:%S') if u.created_at else "",
            status_label,
            _format_dt_human(suspended_until) if status_key == "suspendida" else "",
            int(u.chat_count or 0),
            int(u.message_count or 0),
        ])
    return _build_xlsx_response("usuarios_nexus", "Usuarios", headers, values)


@app.route('/admin/export/login_attempts.xlsx')
@login_required
@admin_required(permission="export_reports")
def admin_export_login_attempts_xlsx():
    rows, _, _ = _admin_security_data(limit=5000)
    headers = ["ID", "IP", "Email", "Intentos", "Primer intento", "Bloqueado hasta"]
    values = []
    for r in rows:
        values.append([
            r.id,
            r.ip or "",
            r.email or "",
            int(r.attempts or 0),
            r.first_attempt_at.strftime('%Y-%m-%d %H:%M:%S') if r.first_attempt_at else "",
            r.blocked_until.strftime('%Y-%m-%d %H:%M:%S') if r.blocked_until else "",
        ])
    return _build_xlsx_response("seguridad_login_attempts", "LoginAttempts", headers, values)


@app.route('/admin/export/auditoria.xlsx')
@login_required
@admin_required(permission="export_reports")
def admin_export_audit_xlsx():
    rows = _admin_logs_for_export(limit=5000)
    enriched = _admin_enrich_logs_rows(rows, utcnow_naive())
    headers = ["Fecha", "Severidad", "Actor email", "Accion", "Usuario objetivo", "IP", "Request ID", "Metodo", "Ruta", "Modulo", "Detalle"]
    values = []
    for item in enriched:
        log_row = item["log_row"]
        actor_user = item["actor_user"]
        values.append([
            log_row.created_at.strftime('%Y-%m-%d %H:%M:%S') if log_row.created_at else "",
            item["severity_label"],
            actor_user.email if actor_user else "",
            log_row.action or "",
            log_row.target_user_id or "",
            log_row.ip or "",
            item["request_id"] or "",
            item["method"] or "",
            item["path"] or "",
            item["module_name"] or "",
            _mask_sensitive_text(log_row.detail or ""),
        ])
    return _build_xlsx_response("auditoria_admin", "Auditoria", headers, values)


@app.route('/admin/export/auditoria.csv')
@login_required
@admin_required(permission="export_reports")
def admin_export_audit_csv():
    rows = _admin_logs_for_export(limit=5000)
    enriched = _admin_enrich_logs_rows(rows, utcnow_naive())

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Fecha", "Severidad", "Actor", "Accion", "Usuario objetivo", "IP", "Request ID", "Metodo", "Ruta", "Modulo", "Detalle"])
    for item in enriched:
        log_row = item["log_row"]
        actor_user = item["actor_user"]
        writer.writerow([
            log_row.created_at.strftime('%Y-%m-%d %H:%M:%S') if log_row.created_at else "",
            item["severity_label"],
            actor_user.email if actor_user else "",
            log_row.action or "",
            log_row.target_user_id or "",
            log_row.ip or "",
            item["request_id"] or "",
            item["method"] or "",
            item["path"] or "",
            item["module_name"] or "",
            _mask_sensitive_text(log_row.detail or ""),
        ])

    data = output.getvalue()
    output.close()
    return Response(
        data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=auditoria_admin.csv"},
    )


@app.route('/admin/export/auditoria.pdf')
@login_required
@admin_required(permission="export_reports")
def admin_export_audit_pdf():
    rows = _admin_logs_for_export(limit=3000)
    enriched = _admin_enrich_logs_rows(rows, utcnow_naive())
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except Exception:
        flash("No se pudo generar PDF. Falta dependencia reportlab.", "error")
        return redirect(url_for("admin_logs_page"))

    def _short(v, size):
        txt = str(v or "")
        return txt if len(txt) <= size else (txt[: size - 3] + "...")

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    y = height - 12 * mm

    c.setTitle("Auditoria administrativa")
    c.setFont("Helvetica-Bold", 12)
    c.drawString(12 * mm, y, "Auditoria administrativa")
    c.setFont("Helvetica", 8)
    c.drawRightString(width - 12 * mm, y, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 7 * mm

    headers = ["Fecha", "Sev", "Accion", "Actor", "IP", "RID", "Detalle"]
    widths = [32, 12, 40, 48, 26, 26, 96]

    def _draw_header(cur_y):
        c.setFont("Helvetica-Bold", 8)
        x = 10 * mm
        for idx, head in enumerate(headers):
            c.drawString(x, cur_y, head)
            x += widths[idx] * mm
        return cur_y - 4.5 * mm

    y = _draw_header(y)
    c.setFont("Helvetica", 7)
    for item in enriched:
        log_row = item["log_row"]
        actor_user = item["actor_user"]
        row_vals = [
            log_row.created_at.strftime('%Y-%m-%d %H:%M:%S') if log_row.created_at else "-",
            item["severity_label"],
            log_row.action or "-",
            actor_user.email if actor_user else "-",
            log_row.ip or "-",
            item["request_id"] or "-",
            _mask_sensitive_text(log_row.detail or "-"),
        ]
        if y < 14 * mm:
            c.showPage()
            y = height - 12 * mm
            y = _draw_header(y)
            c.setFont("Helvetica", 7)
        x = 10 * mm
        for idx, val in enumerate(row_vals):
            limit = 18 if idx in (0, 2, 3, 6) else 12
            c.drawString(x, y, _short(val, limit))
            x += widths[idx] * mm
        y -= 4.2 * mm

    c.save()
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=auditoria_admin.pdf"},
    )


@app.route('/admin/export/usuarios.json')
@login_required
@admin_required(permission="export_reports")
def admin_export_users_json():
    rows = _admin_users_data()
    now_utc = utcnow_naive()
    export_rows = []
    for u in rows:
        status_key, status_label, suspended_until = _user_status_data(u, now_utc)
        export_rows.append([
            u.id,
            u.name or "",
            u.email or "",
            u.created_at.strftime('%Y-%m-%d %H:%M:%S') if u.created_at else "-",
            status_label,
            _format_dt_human(suspended_until) if status_key == "suspendida" else "-",
            int(u.chat_count or 0),
            int(u.message_count or 0),
        ])

    return jsonify({
        "success": True,
        "title": "Usuarios Nexus",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sections": [
            {
                "title": "Usuarios",
                "columns": ["ID", "Nombre", "Email", "Registro", "Estado", "Suspendida hasta", "Chats", "Mensajes"],
                "rows": export_rows,
            }
        ],
    })


@app.route('/admin/export/auditoria.json')
@login_required
@admin_required(permission="export_reports")
def admin_export_audit_json():
    rows = _admin_logs_for_export(limit=5000)
    enriched = _admin_enrich_logs_rows(rows, utcnow_naive())
    return jsonify({
        "success": True,
        "title": "Auditoria administrativa",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sections": [
            {
                "title": "Eventos de auditoria",
                "columns": ["Fecha", "Severidad", "Actor", "Accion", "Usuario objetivo", "IP", "Request ID", "Metodo", "Ruta", "Modulo", "Detalle"],
                "rows": [
                    [
                        item["log_row"].created_at.strftime('%Y-%m-%d %H:%M:%S') if item["log_row"].created_at else "-",
                        item["severity_label"],
                        item["actor_user"].email if item["actor_user"] else "-",
                        item["log_row"].action or "-",
                        item["log_row"].target_user_id or "-",
                        item["log_row"].ip or "-",
                        item["request_id"] or "-",
                        item["method"] or "-",
                        item["path"] or "-",
                        item["module_name"] or "-",
                        _mask_sensitive_text(item["log_row"].detail or "-"),
                    ]
                    for item in enriched
                ],
            }
        ],
    })


@app.route('/admin/export/seguridad.json')
@login_required
@admin_required(permission="export_reports")
def admin_export_security_json():
    login_rows, reset_rows, rate_rows = _admin_security_data(limit=5000)
    return jsonify({
        "success": True,
        "title": "Reporte de seguridad",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sections": [
            {
                "title": "Intentos de login",
                "columns": ["ID", "IP", "Email", "Intentos", "Primer intento", "Bloqueado hasta"],
                "rows": [
                    [
                        r.id,
                        r.ip or "-",
                        r.email or "-",
                        r.attempts or 0,
                        r.first_attempt_at.strftime('%Y-%m-%d %H:%M:%S') if r.first_attempt_at else "-",
                        r.blocked_until.strftime('%Y-%m-%d %H:%M:%S') if r.blocked_until else "-",
                    ]
                    for r in login_rows
                ],
            },
            {
                "title": "Reset por IP",
                "columns": ["IP", "Intentos", "Primer intento", "Bloqueado hasta"],
                "rows": [
                    [
                        r.ip or "-",
                        r.attempts or 0,
                        r.first_attempt_at.strftime('%Y-%m-%d %H:%M:%S') if r.first_attempt_at else "-",
                        r.blocked_until.strftime('%Y-%m-%d %H:%M:%S') if r.blocked_until else "-",
                    ]
                    for r in reset_rows
                ],
            },
            {
                "title": "Rate limits",
                "columns": ["Key", "Count", "Ventana", "Bloqueado hasta"],
                "rows": [
                    [
                        r.key or "-",
                        r.count or 0,
                        r.window_start.strftime('%Y-%m-%d %H:%M:%S') if r.window_start else "-",
                        r.blocked_until.strftime('%Y-%m-%d %H:%M:%S') if r.blocked_until else "-",
                    ]
                    for r in rate_rows
                ],
            },
        ],
    })


@app.route('/admin/grant', methods=['POST'])
@login_required
@admin_required(super_only=True)
def admin_grant():
    email = _normalize_email(request.form.get('email'))
    role = (request.form.get('role') or "admin").strip().lower()
    requested_permissions = set(request.form.getlist('permissions'))

    if not email:
        flash("Debes escribir un correo para asignar permisos.", "error")
        return redirect(url_for('admin_admins_page'))

    user = User.query.filter(func.lower(User.email) == email).first()
    if not user:
        flash("No existe un usuario con ese correo.", "error")
        return redirect(url_for('admin_admins_page'))

    if role not in {"admin", "super_admin"}:
        flash("Rol inválido.", "error")
        return redirect(url_for('admin_admins_page'))

    if role == "super_admin" and not _is_super_admin_email(user.email):
        flash("Solo correos de SUPER_ADMIN_EMAILS pueden ser super admin.", "error")
        return redirect(url_for('admin_admins_page'))

    if role == "super_admin":
        granted_permissions = set(ALL_ADMIN_PERMISSIONS)
    else:
        granted_permissions = {p for p in requested_permissions if p in ALL_ADMIN_PERMISSIONS}
        if not granted_permissions:
            granted_permissions = set(DEFAULT_ADMIN_PERMISSIONS)

    rec = AdminRole.query.filter_by(user_id=user.id).first()
    if not rec:
        rec = AdminRole(
            user_id=user.id,
            role=role,
            permissions_json=_dumps_permissions(granted_permissions),
            is_active=True,
            granted_by_user_id=current_user.id,
        )
        db.session.add(rec)
    else:
        rec.role = role
        rec.permissions_json = _dumps_permissions(granted_permissions)
        rec.is_active = True
        rec.granted_by_user_id = current_user.id
        rec.updated_at = utcnow_naive()

    db.session.commit()
    _add_admin_audit(
        "admin_grant",
        target_user_id=user.id,
        detail=f"email={user.email}; role={role}; permissions={sorted(granted_permissions)}",
    )
    flash("Permisos de administrador actualizados.", "success")
    return redirect(url_for('admin_admins_page'))


@app.route('/admin/revoke/<int:user_id>', methods=['POST'])
@login_required
@admin_required(super_only=True)
def admin_revoke(user_id):
    if user_id == current_user.id:
        flash("No puedes revocar tu propio acceso desde aquí.", "error")
        return redirect(url_for('admin_admins_page'))

    user = db.session.get(User, user_id)
    if not user:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for('admin_admins_page'))

    if _is_super_admin_email(user.email):
        flash("No puedes revocar un super admin por correo protegido.", "error")
        return redirect(url_for('admin_admins_page'))

    rec = AdminRole.query.filter_by(user_id=user_id).first()
    if not rec:
        flash("Ese usuario no tiene rol admin asignado.", "error")
        return redirect(url_for('admin_admins_page'))

    rec.is_active = False
    rec.updated_at = utcnow_naive()
    db.session.commit()
    _add_admin_audit("admin_revoke", target_user_id=user.id, detail=f"email={user.email}")
    flash("Acceso admin revocado.", "success")
    return redirect(url_for('admin_admins_page'))


@app.route('/admin/user_status/<int:user_id>', methods=['POST'])
@login_required
@admin_required(permission="manage_users")
def admin_user_status(user_id):
    user = db.session.get(User, user_id)
    if not user:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for('admin_users_page'))

    if user.id == current_user.id:
        flash("No puedes desactivar tu propia cuenta desde el panel.", "error")
        return redirect(url_for('admin_users_page'))

    if _is_super_admin_email(user.email) and _effective_admin_role(current_user) != "super_admin":
        flash("Solo el super admin puede cambiar esta cuenta.", "error")
        return redirect(url_for('admin_users_page'))

    action = (request.form.get("action") or "").strip().lower()
    if action == "deactivate":
        user.is_active_account = False
        msg = "Cuenta desactivada."
    elif action == "activate":
        user.is_active_account = True
        msg = "Cuenta activada."
    else:
        flash("Accion invalida.", "error")
        return redirect(url_for('admin_users_page'))

    db.session.commit()
    _add_admin_audit("user_status_change", target_user_id=user.id, detail=f"email={user.email}; action={action}")
    flash(msg, "success")
    return redirect(url_for('admin_users_page'))


@app.route('/admin/users_bulk', methods=['POST'])
@login_required
@admin_required(permission="manage_users")
def admin_users_bulk():
    action = (request.form.get("bulk_action") or "").strip().lower()
    raw_ids = request.form.getlist("user_ids")

    user_ids = []
    seen = set()
    for raw in raw_ids:
        try:
            uid = int(raw)
        except (TypeError, ValueError):
            continue
        if uid > 0 and uid not in seen:
            seen.add(uid)
            user_ids.append(uid)

    if not user_ids:
        flash("Selecciona al menos un usuario para aplicar accion masiva.", "error")
        return redirect(url_for('admin_users_page'))

    role = _effective_admin_role(current_user)
    can_export = "export_reports" in _effective_admin_permissions(current_user)

    if action == "export_xlsx":
        if not can_export:
            flash("No tienes permiso para exportar reportes.", "error")
            return redirect(url_for('admin_users_page'))

        selected = {int(uid) for uid in user_ids}
        rows = [u for u in _admin_users_data() if int(u.id) in selected]
        if not rows:
            flash("No se encontraron usuarios validos para exportar.", "error")
            return redirect(url_for('admin_users_page'))

        headers = ["ID", "Nombre", "Email", "Registro", "Estado", "Suspendida hasta", "Chats", "Mensajes"]
        values = []
        now_utc = utcnow_naive()
        for u in rows:
            status_key, status_label, suspended_until = _user_status_data(u, now_utc)
            values.append([
                u.id,
                u.name or "",
                u.email or "",
                u.created_at.strftime('%Y-%m-%d %H:%M:%S') if u.created_at else "",
                status_label,
                _format_dt_human(suspended_until) if status_key == "suspendida" else "",
                int(u.chat_count or 0),
                int(u.message_count or 0),
            ])

        _add_admin_audit(
            "user_bulk_export",
            detail=f"count={len(rows)}; user_ids={','.join(str(uid) for uid in user_ids)}"
        )
        return _build_xlsx_response("usuarios_seleccionados", "UsuariosSeleccionados", headers, values)

    if action not in {"activate", "deactivate"}:
        flash("Accion masiva invalida.", "error")
        return redirect(url_for('admin_users_page'))

    users = User.query.filter(User.id.in_(user_ids)).all()
    changed = 0
    skipped = 0
    target_active = action == "activate"

    for user in users:
        if user.id == current_user.id:
            skipped += 1
            continue
        if _is_super_admin_email(user.email) and role != "super_admin":
            skipped += 1
            continue

        if bool(user.is_active_account) == target_active:
            continue

        user.is_active_account = target_active
        if not target_active:
            user.suspended_until = None
        changed += 1

    db.session.commit()
    _add_admin_audit(
        "user_bulk_status_change",
        detail=(
            f"action={action}; changed={changed}; skipped={skipped}; "
            f"user_ids={','.join(str(uid) for uid in user_ids)}"
        ),
    )

    if changed == 0 and skipped > 0:
        flash("No se aplicaron cambios. Algunos usuarios no se pueden modificar.", "error")
    else:
        suffix = f" (omitidos: {skipped})" if skipped else ""
        label = "activadas" if target_active else "desactivadas"
        flash(f"Cuentas {label}: {changed}{suffix}.", "success")
    return redirect(url_for('admin_users_page'))


@app.route('/admin/user_suspend/<int:user_id>', methods=['POST'])
@login_required
@admin_required(permission="manage_users")
def admin_user_suspend(user_id):
    user = db.session.get(User, user_id)
    if not user:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for('admin_users_page'))

    if user.id == current_user.id:
        flash("No puedes suspender tu propia cuenta desde el panel.", "error")
        return redirect(url_for('admin_users_page'))

    if _is_super_admin_email(user.email) and _effective_admin_role(current_user) != "super_admin":
        flash("Solo el super admin puede cambiar esta cuenta.", "error")
        return redirect(url_for('admin_users_page'))

    action = (request.form.get("action") or "set").strip().lower()
    if action == "clear":
        user.suspended_until = None
        db.session.commit()
        _add_admin_audit("user_unsuspend", target_user_id=user.id, detail=f"email={user.email}")
        flash("Suspension retirada. La cuenta ya puede iniciar sesion.", "success")
        return redirect(url_for('admin_users_page'))

    duration_value = _safe_int(request.form.get("duration_value"), 0)
    duration_unit = (request.form.get("duration_unit") or "").strip().lower()
    if duration_value < 1:
        flash("Debes indicar un tiempo de suspension valido.", "error")
        return redirect(url_for('admin_users_page'))

    if duration_unit == "hours":
        delta = timedelta(hours=duration_value)
    elif duration_unit == "days":
        delta = timedelta(days=duration_value)
    elif duration_unit == "weeks":
        delta = timedelta(weeks=duration_value)
    elif duration_unit == "months":
        delta = timedelta(days=30 * duration_value)
    else:
        flash("Unidad de tiempo invalida.", "error")
        return redirect(url_for('admin_users_page'))

    now_utc = utcnow_naive()
    user.suspended_until = now_utc + delta
    db.session.commit()

    _add_admin_audit(
        "user_suspend",
        target_user_id=user.id,
        detail=(
            f"email={user.email}; duration={duration_value}_{duration_unit}; "
            f"until={_format_dt_human(user.suspended_until)}"
        ),
    )
    flash(f"Cuenta suspendida hasta {_format_dt_human(user.suspended_until)}.", "success")
    return redirect(url_for('admin_users_page'))


@app.route('/admin/user_delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required(super_only=True)
def admin_user_delete(user_id):
    user = db.session.get(User, user_id)
    if not user:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for('admin_users_page'))

    if user.id == current_user.id:
        flash("No puedes eliminar tu propia cuenta.", "error")
        return redirect(url_for('admin_users_page'))

    if _is_super_admin_email(user.email):
        flash("No puedes eliminar un super admin protegido.", "error")
        return redirect(url_for('admin_users_page'))

    convs = Conversation.query.filter_by(user_id=user.id).all()
    conv_ids = [c.id for c in convs]

    if conv_ids:
        SharedConversation.query.filter(SharedConversation.conversation_id.in_(conv_ids)).delete(
            synchronize_session=False
        )
    SharedConversation.query.filter_by(owner_id=user.id).delete(synchronize_session=False)
    SavedMessage.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    AdminRole.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    AdminAuditLog.query.filter(
        (AdminAuditLog.actor_user_id == user.id) | (AdminAuditLog.target_user_id == user.id)
    ).delete(synchronize_session=False)
    Message.query.filter(Message.conversation_id.in_(conv_ids)).delete(synchronize_session=False) if conv_ids else None
    Conversation.query.filter_by(user_id=user.id).delete(synchronize_session=False)

    db.session.delete(user)
    db.session.commit()

    _add_admin_audit("user_delete", target_user_id=user_id, detail=f"user_id={user_id}")
    flash("Usuario eliminado correctamente.", "success")
    return redirect(url_for('admin_users_page'))


@app.route('/new_chat')
@login_required
def new_chat():
    nueva_convo = Conversation(user_id=current_user.id, title="Nuevo Chat")
    db.session.add(nueva_convo)
    db.session.commit()
    return redirect(url_for('home', chat_id=nueva_convo.id))


@app.route('/delete_chat/<int:chat_id>', methods=['POST'])
@login_required
def delete_chat(chat_id):
    chat = db.session.get(Conversation, chat_id)
    if not chat or chat.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Chat no autorizado'}), 403

    try:
        shared_rows = SharedConversation.query.filter_by(conversation_id=chat.id).all()
        shared_tokens = [row.token for row in shared_rows if row.token]
        shared_links_deleted = 0
        viewer_sessions_closed = 0

        if shared_tokens:
            viewer_sessions_closed = SharedViewerPresence.query.filter(
                SharedViewerPresence.token.in_(shared_tokens)
            ).delete(synchronize_session=False)

        shared_links_deleted = SharedConversation.query.filter_by(
            conversation_id=chat.id
        ).delete(synchronize_session=False)

        db.session.delete(chat)
        db.session.commit()
        return jsonify({
            'success': True,
            'shared_links_deleted': int(shared_links_deleted or 0),
            'viewer_sessions_closed': int(viewer_sessions_closed or 0),
        })
    except Exception as exc:
        db.session.rollback()
        log_event(
            "CHAT_DELETE_FAIL",
            user_id=current_user.id,
            chat_id=chat_id,
            err=type(exc).__name__,
        )
        return jsonify({'success': False, 'error': 'No se pudo eliminar este chat ahora.'}), 409


@app.route('/delete_chat_info/<int:chat_id>', methods=['GET'])
@login_required
def delete_chat_info(chat_id):
    chat = db.session.get(Conversation, chat_id)
    if not chat or chat.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Chat no autorizado'}), 403

    shared_rows = SharedConversation.query.filter_by(conversation_id=chat.id).all()
    shared_tokens = [row.token for row in shared_rows if row.token]

    active_viewers = 0
    active_sessions = 0
    if shared_tokens:
        cutoff = utcnow_naive() - timedelta(minutes=3)
        active_rows = SharedViewerPresence.query.filter(
            SharedViewerPresence.token.in_(shared_tokens),
            SharedViewerPresence.last_seen >= cutoff,
        ).all()
        active_viewers = len({
            (row.email or "").strip().lower()
            for row in active_rows
            if (row.email or "").strip()
        })
        active_sessions = len({
            f"{(row.token or '').strip()}::{(row.email or '').strip().lower()}"
            for row in active_rows
            if (row.token or "").strip()
        })

    return jsonify({
        'success': True,
        'chat_id': chat.id,
        'chat_title': chat.title or "Conversacion",
        'has_shared': bool(shared_tokens),
        'shared_links': len(shared_tokens),
        'active_viewers': int(active_viewers),
        'active_sessions': int(active_sessions),
    })


@app.route('/rename_chat/<int:chat_id>', methods=['POST'])
@login_required
def rename_chat(chat_id):
    chat = db.session.get(Conversation, chat_id)
    if not chat or chat.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Chat no autorizado'}), 403

    payload = request.get_json(silent=True) or {}
    title = (payload.get('title') or "").strip()
    if not title:
        return jsonify({'success': False, 'error': 'El título no puede ir vacío'}), 400

    title = _sanitize_text_for_db(" ".join(title.split()))
    if len(title) > 100:
        title = title[:100].rstrip()
    if not title:
        return jsonify({'success': False, 'error': 'Título inválido'}), 400

    if title == (chat.title or "").strip():
        return jsonify({'success': True, 'title': chat.title, 'chat_id': chat.id})

    chat.title = title
    try:
        db.session.commit()
        return jsonify({'success': True, 'title': chat.title, 'chat_id': chat.id})
    except Exception as exc:
        db.session.rollback()
        log_event(
            "CHAT_RENAME_FAIL",
            user_id=current_user.id,
            chat_id=chat_id,
            err=type(exc).__name__,
        )
        return jsonify({'success': False, 'error': 'No se pudo renombrar por ahora.'}), 500


def _bool_flag(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _study_instruction(study_mode: str) -> str:
    mode = (study_mode or "normal").strip().lower()
    if mode == "step":
        return (
            "Responde como tutor académico. Explica paso a paso, sin saltarte pasos, "
            "y al final haz 1 pregunta corta para confirmar si entendió."
        )
    if mode == "hints":
        return "Da únicamente 2-3 pistas y una pregunta guía. No des la solución completa aún."
    if mode == "result":
        return "Da solo el resultado final y una explicación muy breve (1-2 líneas)."
    return ""


IMAGE_MD_RE = re.compile(r'!\[[^\]]*\]\(([^)]+)\)')


def _sanitize_text_for_db(text: str) -> str:
    if text is None:
        return ""
    clean = str(text).replace("\x00", "")
    return "".join(ch for ch in clean if ord(ch) <= 0xFFFF)


def _extract_image_url(md_content: str):
    if not md_content:
        return None
    m = IMAGE_MD_RE.search(md_content)
    return m.group(1).strip() if m else None


def _load_image_from_message_content(md_content: str):
    url = _extract_image_url(md_content or "")
    if not url:
        return None
    try:
        if url.startswith("/static/uploads/"):
            local_path = os.path.join(app.root_path, url.lstrip("/").replace("/", os.sep))
            if os.path.exists(local_path):
                with open(local_path, "rb") as f:
                    return Image.open(BytesIO(f.read()))
            return None
        if url.startswith("http://") or url.startswith("https://"):
            r = requests.get(url, timeout=12)
            if r.ok:
                return Image.open(BytesIO(r.content))
    except Exception:
        return None
    return None


def _build_recent_context(conversation_id: int, limit: int = 6, max_message_id: int = None) -> str:
    q = Message.query.filter_by(conversation_id=conversation_id)
    if max_message_id is not None:
        q = q.filter(Message.id <= max_message_id)

    rows = q.order_by(Message.timestamp.desc(), Message.id.desc()).limit(limit).all()
    rows.reverse()

    lines = []
    for row in rows:
        role = "Usuario" if row.sender == "user" else "Nexus"
        content = (row.content or "").strip()
        if len(content) > 380:
            content = content[:380] + "..."
        lines.append(f"{role}: {content}")
    return "\n".join(lines)


def _mask_key(key: str) -> str:
    if not key:
        return "none"
    return f"...{key[-4:]}"


def _friendly_ai_error(exc: Exception) -> str:
    if isinstance(exc, gexc.ResourceExhausted):
        return "Gemini está temporalmente sin cupo. Intenta de nuevo en unos segundos."
    if isinstance(exc, gexc.DeadlineExceeded):
        return "Gemini tardó demasiado en responder. Intenta con una pregunta más corta."
    if isinstance(exc, gexc.Unauthenticated):
        return "Una clave de Gemini es inválida. Revisa GEMINI_KEYS."
    if isinstance(exc, gexc.PermissionDenied):
        return "Gemini rechazó la solicitud por permisos de la clave."
    if isinstance(exc, gexc.ServiceUnavailable):
        return "Gemini no está disponible en este momento. Intenta de nuevo."
    if isinstance(exc, gexc.InvalidArgument):
        return "Gemini rechazó la solicitud por formato inválido."
    return f"Error de Gemini: {str(exc)[:180]}"


def _resource_query_from_question(question_text: str) -> str:
    q = (question_text or "").strip()
    if not q:
        return ""
    q = IMAGE_MD_RE.sub("", q)
    q = re.sub(r"^Pregunta de [^:]+:\s*", "", q, flags=re.IGNORECASE)
    q = re.sub(r"\s+", " ", q).strip()
    return q[:180]


def _wiki_links(query: str, limit: int = 2):
    if not WIKI_ENABLED or not query:
        return []
    try:
        url = f"https://{WIKI_LANG}.wikipedia.org/w/api.php"
        params = {
            "action": "opensearch",
            "search": query,
            "limit": max(1, min(3, int(limit))),
            "namespace": 0,
            "format": "json",
        }
        r = requests.get(
            url,
            params=params,
            timeout=RESOURCE_HTTP_TIMEOUT_S,
            headers={"User-Agent": "NexusAcademico/1.0"}
        )
        if not r.ok:
            return []
        data = r.json() or []
        titles = data[1] if len(data) > 1 and isinstance(data[1], list) else []
        urls = data[3] if len(data) > 3 and isinstance(data[3], list) else []
        out = []
        for title, link in zip(titles, urls):
            if title and link:
                out.append({"title": str(title).strip(), "url": str(link).strip(), "source": "Wikipedia"})
        return out
    except Exception:
        return []


def _should_include_youtube(query: str) -> bool:
    if not YOUTUBE_ENABLED or not query:
        return False
    q = query.lower()
    kws = [
        "tutorial", "curso", "video", "youtube", "clase",
        "aprender", "paso a paso", "ejercicio", "practica"
    ]
    if any(k in q for k in kws):
        return True
    return random.random() < YOUTUBE_INCLUDE_PROB


def _youtube_links(query: str, limit: int = 2):
    if not YOUTUBE_ENABLED or not query:
        return []
    try:
        params = {
            "part": "snippet",
            "type": "video",
            "q": query,
            "maxResults": max(1, min(3, int(limit))),
            "relevanceLanguage": "es",
            "safeSearch": "moderate",
            "key": YOUTUBE_API_KEY
        }
        r = requests.get(
            "https://www.googleapis.com/youtube/v3/search",
            params=params,
            timeout=RESOURCE_HTTP_TIMEOUT_S
        )
        if not r.ok:
            return []
        data = r.json() or {}
        items = data.get("items") or []
        out = []
        for it in items:
            vid = (((it or {}).get("id") or {}).get("videoId") or "").strip()
            title = (((it or {}).get("snippet") or {}).get("title") or "").strip()
            if vid and title:
                out.append({
                    "title": title,
                    "url": f"https://www.youtube.com/watch?v={vid}",
                    "source": "YouTube"
                })
        return out
    except Exception:
        return []


def _build_learning_links_markdown(question_text: str) -> str:
    query = _resource_query_from_question(question_text)
    if not query:
        return ""

    wiki = _wiki_links(query, limit=2)
    yt = _youtube_links(query, limit=YOUTUBE_MAX_RESULTS) if _should_include_youtube(query) else []

    if not wiki and not yt:
        return ""

    lines = ["### Recursos para profundizar"]
    if wiki and random.random() < WIKI_HINT_PROB:
        note = random.choice([
            "Dato rapido: te dejo una referencia para investigar mas.",
            "Si quieres profundizar, revisa esta fuente de Wikipedia.",
            "Extra de estudio: este enlace te ayuda a ampliar el tema.",
        ])
        lines.append(f"_{note}_")
    for row in wiki:
        lines.append(f"- Wikipedia: [{row['title']}]({row['url']})")
    for row in yt:
        lines.append(f"- Video recomendado: [{row['title']}]({row['url']})")
    return "\n".join(lines)


def _append_learning_links(answer_text: str, question_text: str) -> str:
    base = (answer_text or "").strip()
    if not base:
        return base
    if "### Recursos para profundizar" in base:
        return base
    links = _build_learning_links_markdown(question_text)
    if not links:
        return base
    return f"{base}\n\n---\n{links}"


def _generate_ai_response(*, conversation_id: int, question_text: str, study_mode: str = "normal", img_pil=None, max_message_id: int = None):
    if not LISTA_DE_CLAVES:
        raise RuntimeError("No hay API Key configurada para Gemini")

    context_block = _build_recent_context(
        conversation_id=conversation_id,
        limit=6,
        max_message_id=max_message_id
    )
    mode_block = _study_instruction(study_mode)

    prompt = ""
    if mode_block:
        prompt += mode_block + "\n\n"
    if context_block:
        prompt += f"Contexto reciente de la conversacion:\n{context_block}\n\n"
    prompt += f"Pregunta actual del estudiante:\n{question_text}"
    payload = [img_pil, prompt] if img_pil is not None else [prompt]

    keys = [k for k in LISTA_DE_CLAVES if k]
    random.shuffle(keys)
    max_attempts = max(1, min(len(keys), AI_MAX_KEY_RETRIES if AI_MAX_KEY_RETRIES > 0 else len(keys)))
    attempts = keys[:max_attempts]
    model_candidates = AI_MODEL_CANDIDATES or ["gemini-flash-latest"]

    last_exc = None
    for idx, key in enumerate(attempts, start=1):
        genai.configure(api_key=key)
        for model_name in model_candidates:
            try:
                model = genai.GenerativeModel(
                    model_name=model_name,
                    generation_config=configuracion,
                    system_instruction=instruccion_sistema
                )
                chat_session = model.start_chat(history=[])

                t0 = time.time()
                response = chat_session.send_message(
                    payload,
                    request_options={"timeout": AI_REQUEST_TIMEOUT_S, "retry": None}
                )
                latency_ms = int((time.time() - t0) * 1000)
                raw_text = (response.text or "").replace(r'\hline', '')
                merged_text = _append_learning_links(raw_text, question_text)
                text = _sanitize_text_for_db(merged_text)
                if not text.strip():
                    raise RuntimeError("Gemini devolvio respuesta vacia")

                log_event(
                    "AI_OK",
                    chat_id=conversation_id,
                    attempt=idx,
                    key_mask=_mask_key(key),
                    model=model_name,
                    latency_ms=latency_ms
                )
                return text, latency_ms
            except Exception as e:
                last_exc = e
                log_event(
                    "AI_FAIL",
                    chat_id=conversation_id,
                    attempt=idx,
                    key_mask=_mask_key(key),
                    model=model_name,
                    reason=type(e).__name__,
                    detail=str(e)[:140]
                )
                if isinstance(e, (gexc.InvalidArgument, gexc.PermissionDenied, gexc.Unauthenticated)):
                    raise RuntimeError(_friendly_ai_error(e))
                if isinstance(e, (gexc.NotFound, gexc.ResourceExhausted, gexc.DeadlineExceeded)):
                    continue
                continue

    if last_exc:
        raise RuntimeError(_friendly_ai_error(last_exc))
    raise RuntimeError("Gemini no respondio. Intenta nuevamente.")


def _touch_shared_viewer(token: str, email: str, name: str):
    row = SharedViewerPresence.query.filter_by(token=token, email=email).first()
    now = utcnow_naive()
    if not row:
        row = SharedViewerPresence(token=token, email=email, name=name, last_seen=now)
        db.session.add(row)
    else:
        row.name = name
        row.last_seen = now
    db.session.commit()

    cutoff = now - timedelta(minutes=10)
    stale = SharedViewerPresence.query.filter(
        SharedViewerPresence.token == token,
        SharedViewerPresence.last_seen < cutoff
    ).all()
    for item in stale:
        db.session.delete(item)
    db.session.commit()


def _shared_viewer_count(token: str) -> int:
    cutoff = utcnow_naive() - timedelta(minutes=3)
    rows = SharedViewerPresence.query.filter(
        SharedViewerPresence.token == token,
        SharedViewerPresence.last_seen >= cutoff
    ).all()
    return len({r.email.lower() for r in rows})


@app.route('/share_chat/<int:chat_id>', methods=['POST'])
@login_required
def share_chat(chat_id):
    chat = db.session.get(Conversation, chat_id)
    if not chat or chat.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Chat no encontrado'}), 404

    payload = request.get_json(silent=True) or {}
    permissions = payload.get('permissions') or {}

    read_only = _bool_flag(permissions.get('read_only'), True)
    allow_export = _bool_flag(permissions.get('allow_export'), True)
    allow_copy = _bool_flag(permissions.get('allow_copy'), True)
    allow_feedback = _bool_flag(permissions.get('allow_feedback'), True)
    allow_regenerate = _bool_flag(permissions.get('allow_regenerate'), False) and (not read_only)
    allow_edit = _bool_flag(permissions.get('allow_edit'), False) and (not read_only)

    token = uuid.uuid4().hex + uuid.uuid4().hex[:8]
    shared = SharedConversation(
        token=token,
        conversation_id=chat.id,
        owner_id=current_user.id,
        read_only=read_only,
        allow_export=allow_export,
        allow_copy=allow_copy,
        allow_feedback=allow_feedback,
        allow_regenerate=allow_regenerate,
        allow_edit=allow_edit
    )
    db.session.add(shared)
    db.session.commit()

    share_url = url_for('shared_chat', token=token, _external=True)
    return jsonify({
        'success': True,
        'share_url': share_url,
        'permissions': {
            'read_only': read_only,
            'allow_export': allow_export,
            'allow_copy': allow_copy,
            'allow_feedback': allow_feedback,
            'allow_regenerate': allow_regenerate,
            'allow_edit': allow_edit
        }
    })


@app.route('/shared/<token>', methods=['GET', 'POST'])
def shared_chat(token):
    shared = SharedConversation.query.filter_by(token=token).first()
    if not shared:
        return render_template(
            'shared_unavailable.html',
            title="Enlace no disponible",
            message="Este enlace ya no esta disponible o fue eliminado por el anfitrion.",
            code="404"
        ), 404

    chat = db.session.get(Conversation, shared.conversation_id)
    if not chat:
        return render_template(
            'shared_unavailable.html',
            title="Conversacion eliminada",
            message="La conversacion fue eliminada por el anfitrion.",
            code="410"
        ), 410

    session_email_key = f"shared_email_{token}"
    session_name_key = f"shared_name_{token}"

    if request.method == 'POST':
        email = (request.form.get('email') or "").strip().lower()
        if not EMAIL_RE.match(email):
            return render_template('shared_access.html', token=token, error="Ingresa un correo valido.")

        user = User.query.filter_by(email=email).first()
        if user:
            viewer_name = user.name
        else:
            local_part = email.split('@', 1)[0]
            local_part = local_part.replace('.', ' ').replace('_', ' ').replace('-', ' ')
            local_part = re.sub(r"\s+", " ", local_part).strip()
            viewer_name = local_part.title() if local_part else "Invitado"

        session[session_email_key] = email
        session[session_name_key] = viewer_name
        _touch_shared_viewer(token, email, viewer_name)
        return redirect(url_for('shared_chat', token=token))

    viewer_email = session.get(session_email_key)
    viewer_name = session.get(session_name_key)
    if not viewer_email or not viewer_name:
        return render_template('shared_access.html', token=token, error=None)

    _touch_shared_viewer(token, viewer_email, viewer_name)

    chat_history = (
        Message.query
        .filter_by(conversation_id=chat.id)
        .order_by(Message.timestamp)
        .all()
    )

    permissions = {
        'read_only': bool(shared.read_only),
        'allow_export': bool(shared.allow_export),
        'allow_copy': bool(shared.allow_copy),
        'allow_feedback': bool(shared.allow_feedback),
        'allow_regenerate': bool(shared.allow_regenerate),
        'allow_edit': bool(shared.allow_edit),
    }

    return render_template(
        'shared_chat.html',
        chat_title=chat.title,
        chat_history=chat_history,
        permissions=permissions,
        share_token=shared.token,
        owner_name=chat.owner.name if chat.owner else "Usuario Nexus",
        viewer_name=viewer_name,
        viewer_count=_shared_viewer_count(token)
    )


@app.route('/shared_presence/<token>', methods=['POST'])
def shared_presence(token):
    shared = SharedConversation.query.filter_by(token=token).first()
    if not shared:
        return jsonify({'success': False}), 404

    email = session.get(f"shared_email_{token}")
    name = session.get(f"shared_name_{token}")
    if email and name:
        _touch_shared_viewer(token, email, name)
    return jsonify({'success': True, 'count': _shared_viewer_count(token)})


@app.route('/shared_logout/<token>', methods=['GET', 'POST'])
def shared_logout(token):
    email_key = f"shared_email_{token}"
    name_key = f"shared_name_{token}"
    viewer_email = session.get(email_key)

    try:
        if viewer_email:
            (
                SharedViewerPresence.query
                .filter_by(token=token, email=viewer_email)
                .delete(synchronize_session=False)
            )
            db.session.commit()
    except Exception:
        db.session.rollback()

    session.pop(email_key, None)
    session.pop(name_key, None)
    redirect_url = url_for('login_page')

    if request.method == 'POST':
        return jsonify({'success': True, 'redirect': redirect_url})
    return redirect(redirect_url)


@app.route('/shared_export/<token>', methods=['GET'])
def shared_export(token):
    shared = SharedConversation.query.filter_by(token=token).first()
    if not shared:
        return jsonify({'success': False, 'error': 'Enlace no valido'}), 404
    if not shared.allow_export:
        return jsonify({'success': False, 'error': 'Este enlace no permite exportar'}), 403

    viewer_email = session.get(f"shared_email_{token}")
    viewer_name = session.get(f"shared_name_{token}")
    if not viewer_email or not viewer_name:
        return jsonify({'success': False, 'error': 'Debes validar correo para exportar'}), 401

    chat = db.session.get(Conversation, shared.conversation_id)
    if not chat:
        return jsonify({'success': False, 'error': 'Chat no encontrado'}), 404

    rows = (
        Message.query
        .filter_by(conversation_id=chat.id)
        .order_by(Message.id.asc())
        .all()
    )

    return jsonify({
        'success': True,
        'title': chat.title or 'Conversacion compartida',
        'messages': [
            {
                'id': m.id,
                'sender': m.sender,
                'content': m.content or ''
            }
            for m in rows
        ]
    })


@app.route('/shared_send/<token>', methods=['POST'])
def shared_send(token):
    shared = SharedConversation.query.filter_by(token=token).first()
    if not shared:
        return jsonify({'success': False, 'error': 'Enlace no valido'}), 404
    if shared.read_only or not shared.allow_edit:
        return jsonify({'success': False, 'error': 'Este enlace es de solo lectura'}), 403

    viewer_email = session.get(f"shared_email_{token}")
    viewer_name = session.get(f"shared_name_{token}")
    if not viewer_email or not viewer_name:
        return jsonify({'success': False, 'error': 'Debes validar correo para participar'}), 401

    chat = db.session.get(Conversation, shared.conversation_id)
    if not chat:
        return jsonify({'success': False, 'error': 'Chat no encontrado'}), 404

    message = (request.form.get('message', '') or '').strip()
    study_mode = (request.form.get('study_mode', 'normal') or 'normal').strip().lower()
    image_file = request.files.get('image')

    if not message and not image_file:
        return jsonify({'success': False, 'error': 'Mensaje vacio'}), 400
    if message and len(message) > CHAT_MAX_TEXT_CHARS:
        return jsonify({'success': False, 'error': f'Maximo {CHAT_MAX_TEXT_CHARS} caracteres'}), 400

    image_url = None
    img_pil = None
    if image_file:
        img_bytes = image_file.read()
        if len(img_bytes) > CHAT_MAX_IMAGE_BYTES:
            return jsonify({'success': False, 'error': 'Imagen demasiado grande (8MB max)'}), 400
        image_file.stream.seek(0)
        img_pil = Image.open(BytesIO(img_bytes))

        if CLOUDINARY_URL:
            up = cloudinary.uploader.upload(
                BytesIO(img_bytes),
                folder=f"nexus/{shared.owner_id}/{chat.id}",
                resource_type="image"
            )
            image_url = up.get("secure_url")
        else:
            filename = secure_filename(image_file.filename or "")
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ['.png', '.jpg', '.jpeg', '.gif', '.webp']:
                ext = '.png'
            unique_name = f"shared_{shared.owner_id}_{chat.id}_{int(datetime.now(timezone.utc).timestamp())}_{random.randint(1000,9999)}{ext}"
            image_path = os.path.join(UPLOAD_DIR, unique_name)
            with open(image_path, "wb") as f:
                f.write(img_bytes)
            image_url = f"/static/uploads/{unique_name}"

    label = f"({viewer_name}) "
    if image_url:
        img_md = f"![Imagen enviada]({image_url})"
        body = f"{img_md}\n\n{label}{message}" if message else f"{img_md}\n\n{label}"
    else:
        body = f"{label}{message}"

    user_msg = Message(
        content=_sanitize_text_for_db(body),
        sender='user',
        conversation_id=chat.id
    )
    user_msg.has_image = bool(image_url)
    db.session.add(user_msg)
    db.session.commit()

    question_text = message if message else "Analiza esta imagen y explica que ves."
    if viewer_name:
        question_text = f"Pregunta de {viewer_name}: {question_text}"

    try:
        bot_text, latency_ms = _generate_ai_response(
            conversation_id=chat.id,
            question_text=question_text,
            study_mode=study_mode,
            img_pil=img_pil,
            max_message_id=user_msg.id
        )
        bot_msg = Message(
            content=_sanitize_text_for_db(bot_text),
            sender='bot',
            conversation_id=chat.id
        )
        db.session.add(bot_msg)
        db.session.commit()

        log_event(
            "SHARED_CHAT_SENT",
            chat_id=chat.id,
            owner_id=shared.owner_id,
            viewer=viewer_email,
            latency_ms=latency_ms
        )

        return jsonify({
            'success': True,
            'response': bot_msg.content,
            'user_message_id': user_msg.id,
            'bot_message_id': bot_msg.id
        })
    except Exception as e:
        logger.exception("SHARED_SEND_ERROR chat_id=%s viewer=%s", chat.id if chat else None, viewer_email)
        return jsonify({'success': False, 'error': str(e) or 'No se pudo enviar'}), 500


@app.route('/shared_regenerate/<token>', methods=['POST'])
def shared_regenerate(token):
    shared = SharedConversation.query.filter_by(token=token).first()
    if not shared:
        return jsonify({'success': False, 'error': 'Enlace no valido'}), 404
    if shared.read_only or not shared.allow_regenerate:
        return jsonify({'success': False, 'error': 'Este enlace no permite regenerar'}), 403

    viewer_email = session.get(f"shared_email_{token}")
    if not viewer_email:
        return jsonify({'success': False, 'error': 'Debes validar correo para participar'}), 401

    payload = request.get_json(silent=True) or {}
    bot_message_id = payload.get('bot_message_id')
    if not bot_message_id:
        return jsonify({'success': False, 'error': 'Falta bot_message_id'}), 400

    try:
        bot_message_id = int(bot_message_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'ID invalido'}), 400

    chat_id = shared.conversation_id
    bot_msg = db.session.get(Message, bot_message_id)
    if not bot_msg or bot_msg.conversation_id != chat_id or bot_msg.sender != 'bot':
        return jsonify({'success': False, 'error': 'Mensaje bot invalido'}), 400

    user_msg = (
        Message.query
        .filter(
            Message.conversation_id == chat_id,
            Message.sender == 'user',
            Message.id < bot_msg.id
        )
        .order_by(Message.id.desc())
        .first()
    )
    if not user_msg:
        return jsonify({'success': False, 'error': 'No se encontro mensaje previo'}), 400

    question_text = (user_msg.content or "").strip()
    image_url = _extract_image_url(question_text) if user_msg.has_image else None
    if image_url:
        question_text = IMAGE_MD_RE.sub("", question_text).strip()
    if not question_text:
        question_text = "Analiza de nuevo esta imagen y explica con claridad."

    try:
        bot_text, latency_ms = _generate_ai_response(
            conversation_id=chat_id,
            question_text=question_text,
            study_mode=(payload.get('study_mode') or 'normal'),
            img_pil=_load_image_from_message_content(user_msg.content) if user_msg.has_image else None,
            max_message_id=user_msg.id
        )
        bot_msg.content = _sanitize_text_for_db(bot_text)
        db.session.commit()

        log_event(
            "SHARED_CHAT_REGENERATE",
            chat_id=chat_id,
            owner_id=shared.owner_id,
            viewer=viewer_email,
            latency_ms=latency_ms
        )
        return jsonify({'success': True, 'response': bot_msg.content, 'bot_message_id': bot_msg.id})
    except Exception as e:
        logger.exception("SHARED_REGENERATE_ERROR chat_id=%s viewer=%s", chat_id, viewer_email)
        return jsonify({'success': False, 'error': str(e) or 'No se pudo regenerar'}), 500


SAVED_MAX_ITEMS_PER_USER = 500


def _parse_client_iso_to_naive_utc(value: str):
    raw = (value or "").strip()
    if not raw:
        return utcnow_naive()
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return to_naive_utc(dt)
    except Exception:
        return utcnow_naive()


def _prune_saved_messages(user_id: int, keep: int = SAVED_MAX_ITEMS_PER_USER):
    rows = (
        SavedMessage.query
        .filter_by(user_id=user_id)
        .order_by(SavedMessage.created_at.desc(), SavedMessage.id.desc())
        .all()
    )
    if len(rows) <= keep:
        return
    for item in rows[keep:]:
        db.session.delete(item)


@app.route('/saved_messages', methods=['GET'])
@login_required
def list_saved_messages():
    rows = (
        SavedMessage.query
        .filter_by(user_id=current_user.id)
        .order_by(SavedMessage.created_at.desc(), SavedMessage.id.desc())
        .all()
    )
    return jsonify({
        'success': True,
        'items': [
            {
                'id': row.id,
                'text': row.content or '',
                'ts': f"{row.created_at.isoformat()}Z" if row.created_at else None
            }
            for row in rows
        ]
    })


@app.route('/saved_messages', methods=['POST'])
@login_required
def create_saved_message():
    payload = request.get_json(silent=True) or {}
    raw_text = (payload.get('text') or '').strip()
    if not raw_text:
        return jsonify({'success': False, 'error': 'No hay contenido para guardar'}), 400

    clean_text = _sanitize_text_for_db(raw_text)
    if not clean_text.strip():
        return jsonify({'success': False, 'error': 'Contenido invalido'}), 400
    if len(clean_text) > 120000:
        return jsonify({'success': False, 'error': 'Contenido demasiado largo'}), 400

    now = utcnow_naive()
    latest_same = (
        SavedMessage.query
        .filter_by(user_id=current_user.id, content=clean_text)
        .order_by(SavedMessage.created_at.desc(), SavedMessage.id.desc())
        .first()
    )
    if latest_same and latest_same.created_at and (now - latest_same.created_at).total_seconds() <= 15:
        return jsonify({
            'success': True,
            'item': {
                'id': latest_same.id,
                'text': latest_same.content,
                'ts': f"{latest_same.created_at.isoformat()}Z"
            },
            'dedup': True
        })

    row = SavedMessage(content=clean_text, user_id=current_user.id, created_at=now)
    db.session.add(row)
    _prune_saved_messages(current_user.id)
    db.session.commit()
    return jsonify({
        'success': True,
        'item': {
            'id': row.id,
            'text': row.content,
            'ts': f"{row.created_at.isoformat()}Z"
        }
    })


@app.route('/saved_messages/sync', methods=['POST'])
@login_required
def sync_saved_messages():
    payload = request.get_json(silent=True) or {}
    items = payload.get('items') or []
    if not isinstance(items, list):
        return jsonify({'success': False, 'error': 'Formato invalido'}), 400

    inserted = 0
    for item in items[:200]:
        if not isinstance(item, dict):
            continue
        text = _sanitize_text_for_db((item.get('text') or '').strip())
        if not text:
            continue
        exists = SavedMessage.query.filter_by(user_id=current_user.id, content=text).first()
        if exists:
            continue
        row = SavedMessage(
            content=text,
            user_id=current_user.id,
            created_at=_parse_client_iso_to_naive_utc(item.get('ts') or '')
        )
        db.session.add(row)
        inserted += 1

    _prune_saved_messages(current_user.id)
    db.session.commit()
    return jsonify({'success': True, 'inserted': inserted})


@app.route('/saved_messages/<int:item_id>', methods=['DELETE'])
@login_required
def delete_saved_message(item_id):
    row = SavedMessage.query.filter_by(id=item_id, user_id=current_user.id).first()
    if not row:
        return jsonify({'success': False, 'error': 'Guardado no encontrado'}), 404
    db.session.delete(row)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/saved_messages', methods=['DELETE'])
@login_required
def clear_saved_messages():
    (
        SavedMessage.query
        .filter_by(user_id=current_user.id)
        .delete(synchronize_session=False)
    )
    db.session.commit()
    return jsonify({'success': True})


@app.route('/edit_and_resend', methods=['POST'])
@login_required
def edit_and_resend():
    payload = request.get_json(silent=True) or {}
    chat_id = payload.get('chat_id')
    message_id = payload.get('message_id')
    new_text = (payload.get('message') or '').strip()
    study_mode = (payload.get('study_mode') or 'normal').strip().lower()

    if not chat_id or not message_id or not new_text:
        return jsonify({'success': False, 'error': 'Datos incompletos'}), 400

    try:
        chat_id = int(chat_id)
        message_id = int(message_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'IDs inválidos'}), 400

    if len(new_text) > CHAT_MAX_TEXT_CHARS:
        return jsonify({'success': False, 'error': f'Máximo {CHAT_MAX_TEXT_CHARS} caracteres'}), 400

    chat = db.session.get(Conversation, chat_id)
    if not chat or chat.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Chat no autorizado'}), 403

    msg = db.session.get(Message, message_id)
    if not msg or msg.conversation_id != chat.id or msg.sender != 'user':
        return jsonify({'success': False, 'error': 'Mensaje inválido'}), 400

    last_user = (
        Message.query
        .filter_by(conversation_id=chat.id, sender='user')
        .order_by(Message.timestamp.desc(), Message.id.desc())
        .first()
    )
    if not last_user or last_user.id != msg.id:
        return jsonify({'success': False, 'error': 'Solo puedes editar el último mensaje de usuario'}), 400

    image_url = _extract_image_url(msg.content) if msg.has_image else None
    if msg.has_image and image_url:
        img_md = f"![Imagen enviada]({image_url})"
        msg.content = f"{img_md}\n\n{new_text}" if new_text else img_md
    else:
        msg.content = new_text
    msg.content = _sanitize_text_for_db(msg.content)

    tail = Message.query.filter(
        Message.conversation_id == chat.id,
        Message.id > msg.id
    ).all()
    for item in tail:
        db.session.delete(item)
    db.session.commit()

    try:
        question_text = new_text or "Analiza de nuevo esta imagen y explica con claridad."
        img_for_ai = _load_image_from_message_content(msg.content) if msg.has_image else None
        texto_limpio, latency_ms = _generate_ai_response(
            conversation_id=chat.id,
            question_text=question_text,
            study_mode=study_mode,
            img_pil=img_for_ai,
            max_message_id=msg.id
        )

        bot_msg = Message(
            content=_sanitize_text_for_db(texto_limpio),
            sender='bot',
            conversation_id=chat.id
        )
        db.session.add(bot_msg)
        db.session.commit()

        log_event(
            "CHAT_EDIT_RESEND",
            user_id=current_user.id,
            chat_id=chat.id,
            msg_id=msg.id,
            latency_ms=latency_ms
        )

        return jsonify({
            'success': True,
            'response': texto_limpio,
            'chat_id': chat.id,
            'user_message_id': msg.id,
            'bot_message_id': bot_msg.id
        })
    except Exception as e:
        logger.exception("EDIT_AND_RESEND_ERROR user_id=%s chat_id=%s", current_user.id, chat.id if chat else None)
        return jsonify({'success': False, 'error': str(e) or 'No se pudo regenerar la respuesta'}), 500


@app.route('/regenerate_response', methods=['POST'])
@login_required
def regenerate_response():
    payload = request.get_json(silent=True) or {}
    chat_id = payload.get('chat_id')
    bot_message_id = payload.get('bot_message_id')

    if not chat_id or not bot_message_id:
        return jsonify({'success': False, 'error': 'Datos incompletos'}), 400

    try:
        chat_id = int(chat_id)
        bot_message_id = int(bot_message_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'IDs invalidos'}), 400

    chat = db.session.get(Conversation, chat_id)
    if not chat or chat.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Chat no autorizado'}), 403

    bot_msg = db.session.get(Message, bot_message_id)
    if not bot_msg or bot_msg.conversation_id != chat_id or bot_msg.sender != 'bot':
        return jsonify({'success': False, 'error': 'Mensaje bot invalido'}), 400

    user_msg = (
        Message.query
        .filter(
            Message.conversation_id == chat_id,
            Message.sender == 'user',
            Message.id < bot_msg.id
        )
        .order_by(Message.id.desc())
        .first()
    )
    if not user_msg:
        return jsonify({'success': False, 'error': 'No se encontro mensaje de usuario previo'}), 400

    try:
        question_text = (user_msg.content or "").strip()
        image_url = _extract_image_url(question_text) if user_msg.has_image else None
        if image_url:
            question_text = IMAGE_MD_RE.sub("", question_text).strip()
        if not question_text:
            question_text = "Analiza de nuevo esta imagen y explica con claridad."

        img_for_ai = _load_image_from_message_content(user_msg.content) if user_msg.has_image else None
        texto_limpio, latency_ms = _generate_ai_response(
            conversation_id=chat_id,
            question_text=question_text,
            study_mode=(payload.get('study_mode') or 'normal'),
            img_pil=img_for_ai,
            max_message_id=user_msg.id
        )

        bot_msg.content = _sanitize_text_for_db(texto_limpio)
        db.session.commit()

        log_event(
            "CHAT_REGENERATE",
            user_id=current_user.id,
            chat_id=chat_id,
            user_msg_id=user_msg.id,
            bot_msg_id=bot_msg.id,
            latency_ms=latency_ms
        )

        return jsonify({
            'success': True,
            'response': bot_msg.content,
            'bot_message_id': bot_msg.id,
            'user_message_id': user_msg.id
        })
    except Exception as e:
        logger.exception("REGENERATE_RESPONSE_ERROR user_id=%s chat_id=%s", current_user.id, chat_id)
        return jsonify({'success': False, 'error': str(e) or 'No se pudo regenerar'}), 500


@app.route('/chat', methods=['POST'])
@login_required
def chat():
    # Valida que existan keys (la selección/reintento se maneja en _generate_ai_response)
    if not LISTA_DE_CLAVES:
        print("ERROR: GEMINI_KEYS vacio o no cargado")
        msg = "No hay API Key configurada para Gemini. Revisa GEMINI_KEYS."
        return jsonify({'success': False, 'error': msg, 'response': msg}), 500
    # =========================================================
    # Fase 2 — Paso 2.6: Rate limit para /chat (protege costos)
    # =========================================================
    ip = get_client_ip()
    uid = current_user.id if current_user.is_authenticated else 0
    
    rl_ok, rl_wait = rate_limit_check(
        key=_rl_key("chat", ip, uid),
        max_count=CHAT_RL_MAX,
        window_seconds=CHAT_RL_WINDOW_S,
        block_seconds=CHAT_RL_BLOCK_S
    )
    if not rl_ok:
        return jsonify({'response': f"Demasiados mensajes. Espera {rl_wait} segundos e intenta de nuevo."}), 429
    # =========================================================

    mensaje_usuario = request.form.get('message', '')

    # =========================================================
    # Fase 3 — Modo de estudio (Paso 4)
    # =========================================================
    study_mode = (request.form.get('study_mode', 'normal') or 'normal').strip().lower()

    # =========================================================
    # Fase 2 — Paso 2.6: Límite de caracteres en texto
    # =========================================================
    if mensaje_usuario and len(mensaje_usuario) > CHAT_MAX_TEXT_CHARS:
        return jsonify({'response': f"Tu mensaje es muy largo. Máximo {CHAT_MAX_TEXT_CHARS} caracteres."}), 400
    # =========================================================

    chat_id = request.form.get('chat_id')
    imagen_archivo = request.files.get('image')

    if not mensaje_usuario and not imagen_archivo:
        return jsonify({'response': '...'})

    if not chat_id or chat_id == 'None' or chat_id == '':
        nueva_convo = Conversation(user_id=current_user.id, title="Nuevo Chat")
        db.session.add(nueva_convo)
        db.session.commit()
        chat_id = nueva_convo.id
    else:
        chat_id = int(chat_id)

    try:
        image_url = None
        img_pil = None
        img_bytes = None

        if imagen_archivo:
            img_bytes = imagen_archivo.read()

            # =========================================================
            # Fase 2 — Paso 2.6: Límite de tamaño en imágenes
            # =========================================================
            if len(img_bytes) > CHAT_MAX_IMAGE_BYTES:
                return jsonify({'response': "La imagen es demasiado grande. Máximo 8MB."}), 400
            # =========================================================
            
            imagen_archivo.stream.seek(0)
            img_pil = Image.open(BytesIO(img_bytes))

            if CLOUDINARY_URL:
                up = cloudinary.uploader.upload(
                    BytesIO(img_bytes),
                    folder=f"nexus/{current_user.id}/{chat_id}",
                    resource_type="image"
                )
                image_url = up.get("secure_url")
            else:
                filename = secure_filename(imagen_archivo.filename or "")
                ext = os.path.splitext(filename)[1].lower()
                if ext not in ['.png', '.jpg', '.jpeg', '.gif', '.webp']:
                    ext = '.png'
                unique_name = f"{current_user.id}_{chat_id}_{int(datetime.now(timezone.utc).timestamp())}_{random.randint(1000,9999)}{ext}"  # FIX: evitar datetime.utcnow() (deprecated)
                image_path = os.path.join(UPLOAD_DIR, unique_name)
                with open(image_path, "wb") as f:
                    f.write(img_bytes)
                image_url = f"/static/uploads/{unique_name}"

        if image_url:
            img_md = f"![Imagen enviada]({image_url})"
            contenido_msg = f"{img_md}\n\n{mensaje_usuario}" if mensaje_usuario else img_md
        else:
            contenido_msg = mensaje_usuario

        msg_db = Message(
            content=_sanitize_text_for_db(contenido_msg if contenido_msg else "[Imagen enviada]"),
            sender='user',
            conversation_id=chat_id
        )
        msg_db.has_image = bool(image_url)
        db.session.add(msg_db)
        db.session.commit()

        question_text = mensaje_usuario if mensaje_usuario else "Analiza esta imagen y explica que ves."
        texto_limpio, latency_ms = _generate_ai_response(
            conversation_id=chat_id,
            question_text=question_text,
            study_mode=study_mode,
            img_pil=img_pil,
            max_message_id=msg_db.id
        )

        log_event(
            "CHAT_SENT",
            user_id=current_user.id,
            chat_id=chat_id,
            has_image=bool(image_url),
            text_len=len(mensaje_usuario or ""),
            latency_ms=latency_ms
        )

        # FIX: SQLAlchemy 2.0 reemplaza Query.get() por session.get()
        convo = db.session.get(Conversation, chat_id)
        new_title = None
        if convo and convo.title == "Nuevo Chat":
            titulo_base = mensaje_usuario if mensaje_usuario else "Imagen Analizada"
            convo.title = _sanitize_text_for_db(" ".join(titulo_base.split()[:4]) + "...")[:100]
            db.session.commit()
            new_title = convo.title

        bot_msg_db = Message(
            content=_sanitize_text_for_db(texto_limpio),
            sender='bot',
            conversation_id=chat_id
        )
        db.session.add(bot_msg_db)
        db.session.commit()

        return jsonify({
            'response': texto_limpio,
            'chat_id': chat_id,
            'new_title': new_title,
            'user_message_id': msg_db.id,
            'bot_message_id': bot_msg_db.id
        })

    except Exception as e:
        logger.exception(
            "CHAT_ERROR user_id=%s chat_id=%s err=%s",
            current_user.id if current_user.is_authenticated else None,
            chat_id,
            type(e).__name__
        )
        err_msg = str(e).strip() or "Tuve un problema técnico procesando eso. Intenta de nuevo."
        links_md = _build_learning_links_markdown(mensaje_usuario or "")
        err_body = f"Nexus no pudo responder: {err_msg}"
        if links_md:
            err_body = f"{err_body}\n\n---\n{links_md}"
        bot_message_id = None
        try:
            cid = int(chat_id) if chat_id else None
            if cid:
                bot_err = Message(
                    content=_sanitize_text_for_db(err_body),
                    sender='bot',
                    conversation_id=cid
                )
                db.session.add(bot_err)
                db.session.commit()
                bot_message_id = bot_err.id
        except Exception:
            db.session.rollback()
        return jsonify({'success': False, 'error': err_msg, 'response': err_body, 'bot_message_id': bot_message_id}), 502

print("GEMINI_KEYS exist?:", bool(os.getenv("GEMINI_KEYS")))
print("GEMINI_KEYS count:", len(LISTA_DE_CLAVES))

# =========================================================
# 15) EJECUCIÓN (Configurado para Railway/Producción)
# =========================================================
if __name__ == '__main__':
    # Railway asigna un puerto dinámico en la variable de entorno 'PORT'
    # Si no existe (ej. corriendo local), usará el 5000 por defecto
    port = int(os.environ.get("PORT", 5000))
    
    # IMPORTANTE: host='0.0.0.0' permite que el contenedor reciba tráfico externo
    app.run(host='0.0.0.0', port=port)
